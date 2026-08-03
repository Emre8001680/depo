import io
import base64
import hashlib
import json
from datetime import datetime, date, timedelta
from zoneinfo import ZoneInfo
import pandas as pd
import streamlit as st
import streamlit.components.v1 as components
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from supabase import create_client, Client

# Streamlit gereği sayfa yapılandırması ilk Streamlit komutu olmalıdır.
st.set_page_config(page_title="Yalçın Marketler Zinciri - Manav Portalı", page_icon="🥭", layout="wide")

# -------------------------------------------------------------
# 🌐 SUPABASE BAĞLANTI BİLGİLERİ
# -------------------------------------------------------------
try:
    SUPABASE_URL = st.secrets["SUPABASE_URL"]
    SUPABASE_KEY = st.secrets["SUPABASE_KEY"]
except KeyError:
    st.error("Sistem bağlantı ayarları eksik: SUPABASE_URL ve SUPABASE_KEY tanımlanmalıdır.")
    st.stop()

@st.cache_resource
def init_supabase() -> Client:
    return create_client(SUPABASE_URL, SUPABASE_KEY)

try:
    supabase = init_supabase()
except Exception as e:
    st.error("Supabase bağlantısı kurulamadı. Lütfen bağlantı bilgilerini kontrol edin.")
    st.exception(e)
    st.stop()

# -------------------------------------------------------------
# 🎨 CSS DÜZENLEMELERİ
# -------------------------------------------------------------
st.markdown("""
    <style>
        #MainMenu {visibility: hidden !important;}
        footer {visibility: hidden !important;}
        header {visibility: hidden !important;}
        [data-testid="stHeader"] {display: none !important;}
        [data-testid="stSidebar"] {display: none !important; }
        
        .block-container {
            padding-top: 2rem !important;
            padding-bottom: 2rem !important;
        }

        @keyframes fadeInZoom {
            0% { opacity: 0; transform: scale(0.9); }
            100% { opacity: 1; transform: scale(1); }
        }

        .logo-card-container {
            animation: fadeInZoom 1s ease-out forwards;
            background-color: #ffffff !important;
            border-radius: 20px;
            padding: 30px 20px;
            box-shadow: 0 10px 25px rgba(0, 0, 0, 0.15);
            max-width: 460px;
            margin: 0 auto 20px auto;
            display: flex;
            justify-content: center;
            align-items: center;
        }

        .animated-logo {
            max-width: 100%;
            height: auto;
            display: block;
        }

        .welcome-title {
            text-align: center;
            font-size: 26px;
            font-weight: 800;
            letter-spacing: 1px;
            color: var(--text-color) !important;
            margin-top: 15px;
            margin-bottom: 5px;
        }
        
        .welcome-sub {
            text-align: center;
            font-size: 15px;
            color: var(--text-color) !important;
            opacity: 0.75;
            margin-bottom: 25px;
        }
    </style>
""", unsafe_allow_html=True)

SUBE_LISTESI = [
    "Raufbey", "Metin Tamer", "Hacı Osmanlı", "Salı Yolu", "Kadiri Yolu", 
    "Nahır Yolu", "Eyup Sultan", "Bulvar", "Düziçi Çarşı", "Aşiyan", "Zeytinlik"
]

SUBE_SIFRELERI = {
    "Raufbey": "1001", "Metin Tamer": "1002", "Hacı Osmanlı": "1003",
    "Salı Yolu": "1004", "Kadiri Yolu": "1005", "Nahır Yolu": "1006",
    "Eyup Sultan": "1007", "Bulvar": "1008", "Düziçi Çarşı": "1009",
    "Aşiyan": "1010", "Zeytinlik": "1011"
}

HAL_SIFRESI = st.secrets.get("HAL_PASSWORD")
YONETICI_SIFRESI = st.secrets.get("ADMIN_PASSWORD")

# Canlı ortamda yönetici şifrelerinin mutlaka Streamlit Secrets üzerinden gelmesi gerekir.
if not HAL_SIFRESI or not YONETICI_SIFRESI:
    st.error("Sistem güvenlik ayarları eksik: HAL_PASSWORD ve ADMIN_PASSWORD tanımlanmalıdır.")
    st.stop()

# Türkiye saati.
ISTANBUL_TZ = ZoneInfo("Europe/Istanbul")

URUNLER = [{'KODU': '053016', 'ADI': 'MNV.ACI DOLMALIK'},
 {'KODU': '09950', 'ADI': 'MNV.ACUR'},
 {'KODU': '09857', 'ADI': 'MNV.ALA KARPUZ'},
 {'KODU': '00015264', 'ADI': 'MNV.ANANAS'},
 {'KODU': '08385', 'ADI': 'MNV.ARMUT'},
 {'KODU': '84', 'ADI': 'MNV.AVOKADO ADET'},
 {'KODU': '09922', 'ADI': 'MNV.AYVA'},
 {'KODU': '058418', 'ADI': 'MNV.BAMYA'},
 {'KODU': '09921', 'ADI': 'MNV.BARBUNYA'},
 {'KODU': '09952', 'ADI': 'MNV.BEYAZ LAHANA ADET'},
 {'KODU': '055710', 'ADI': 'MNV.BEYAZ SOGAN'},
 {'KODU': '04587', 'ADI': 'MNV.BEYAZ TURP'},
 {'KODU': '01248', 'ADI': 'MNV.BEYAZ UZUM'},
 {'KODU': '056484', 'ADI': 'MNV.BEZELYE'},
 {'KODU': '8684523730129', 'ADI': 'MNV.BOGURTLEN 125GR'},
 {'KODU': '01257', 'ADI': 'MNV.BROKOLI'},
 {'KODU': '09953', 'ADI': 'MNV.BRUKSEL PAKET LAHANA ADET'},
 {'KODU': '09965', 'ADI': 'MNV.BURSA DOMATES'},
 {'KODU': 'B.2901020', 'ADI': 'MNV.BURSA SEFTALI'},
 {'KODU': '03718', 'ADI': 'MNV.CAGLA'},
 {'KODU': '05695', 'ADI': 'MNV.CARLISTON BIBER'},
 {'KODU': '09859', 'ADI': 'MNV.CEKIRDEKSIZ KARPUZ'},
 {'KODU': '04239', 'ADI': 'MNV.CEKIRDEKSIZ UZUM'},
 {'KODU': '09911', 'ADI': 'MNV.CERI DOMATES'},
 {'KODU': '01127', 'ADI': 'MNV.CILEK'},
 {'KODU': '00001922', 'ADI': 'MNV.DERE OTU'},
 {'KODU': '09949', 'ADI': 'MNV.DEVECI ARMUT'},
 {'KODU': '05485', 'ADI': 'MNV.DOLMALIK BIBER'},
 {'KODU': 'B.2801083', 'ADI': 'MNV.EJDER MEYVESI ADET'},
 {'KODU': '07704', 'ADI': 'MNV.ELMA ARJANTIN'},
 {'KODU': '07703', 'ADI': 'MNV.ELMA GOLDEN'},
 {'KODU': '07706', 'ADI': 'MNV.ELMA PINK LADY'},
 {'KODU': '07701', 'ADI': 'MNV.ELMA STARKING'},
 {'KODU': 'B.2801088', 'ADI': 'MNV.ENGINAR ADET'},
 {'KODU': '09966', 'ADI': 'MNV.ERIK KG'},
 {'KODU': '08383', 'ADI': 'MNV.ERIK PAKET'},
 {'KODU': '06108', 'ADI': 'MNV.FASULYE YESIL'},
 {'KODU': '120', 'ADI': 'MNV.FIRIK MISIR'},
 {'KODU': '052996', 'ADI': 'MNV.FIRIK SOGAN'},
 {'KODU': '81', 'ADI': 'MNV.GOBEKLI MARUL'},
 {'KODU': '09396', 'ADI': 'MNV.GREYFURT'},
 {'KODU': '39', 'ADI': 'MNV.HINDISTAN CEVIZI'},
 {'KODU': 'B.2901061', 'ADI': 'MNV.HURMA SARI'},
 {'KODU': '09901', 'ADI': 'MNV.HURMA SIYAH'},
 {'KODU': '09941', 'ADI': 'MNV.ISPANAK'},
 {'KODU': '024179', 'ADI': 'MNV.ITALYAN ERIK'},
 {'KODU': '053743', 'ADI': 'MNV.ITHAL MUZ'},
 {'KODU': '01785', 'ADI': 'MNV.JALAPENO BIBER'},
 {'KODU': '05773', 'ADI': 'MNV.KABAK BEYAZ'},
 {'KODU': '06374', 'ADI': 'MNV.KABAK SIYAH'},
 {'KODU': '053736', 'ADI': 'MNV.KAMKAT'},
 {'KODU': '09858', 'ADI': 'MNV.KARA KARPUZ'},
 {'KODU': '07603', 'ADI': 'MNV.KARA LAHANA'},
 {'KODU': '09356', 'ADI': 'MNV.KARNABAHAR'},
 {'KODU': '07451', 'ADI': 'MNV.KAVUN ANKARA'},
 {'KODU': 'B.2901023', 'ADI': 'MNV.KAVURMALIK SOGAN'},
 {'KODU': '053054', 'ADI': 'MNV.KAYISI'},
 {'KODU': '05790', 'ADI': 'MNV.KEMER PATLICAN'},
 {'KODU': '055603', 'ADI': 'MNV.KEREVIZ'},
 {'KODU': '052826', 'ADI': 'MNV.KESTANE'},
 {'KODU': '09935', 'ADI': 'MNV.KIRAZ'},
 {'KODU': '01151', 'ADI': 'MNV.KIRMIZI KAPYA BIBER'},
 {'KODU': '056063', 'ADI': 'MNV.KIRMIZI PANCAR'},
 {'KODU': '01153', 'ADI': 'MNV.KIRMIZI SILI BIBER'},
 {'KODU': '06375', 'ADI': 'MNV.KIRMIZI SOGAN'},
 {'KODU': '09900', 'ADI': 'MNV.KIRMIZI TURP'},
 {'KODU': '09398', 'ADI': 'MNV.KIVI'},
 {'KODU': '76', 'ADI': 'MNV.KIVIRCIK MARUL'},
 {'KODU': 'B.2901017', 'ADI': 'MNV.KOZLEMELIK PATLICAN'},
 {'KODU': '053197', 'ADI': 'MNV.KURU SARIMSAK'},
 {'KODU': '09934', 'ADI': 'MNV.LIMON'},
 {'KODU': 'B.2901019', 'ADI': 'MNV.LUX PATATES'},
 {'KODU': '09997', 'ADI': 'MNV.LUX SALATALIK'},
 {'KODU': '053056', 'ADI': 'MNV.MALATYA KAYISI'},
 {'KODU': '09967', 'ADI': 'MNV.MANDALINA'},
 {'KODU': '051279', 'ADI': 'MNV.MANGO ADET'},
 {'KODU': '8699211220011', 'ADI': 'MNV.MANTAR PAKET'},
 {'KODU': '01487', 'ADI': 'MNV.MARGARIT ARMUT'},
 {'KODU': '69', 'ADI': 'MNV.MARUL'},
 {'KODU': '70', 'ADI': 'MNV.MAYDANOZ'},
 {'KODU': 'MNV.017376', 'ADI': 'MNV.MURDUM ERIK KG'},
 {'KODU': '014146', 'ADI': 'MNV.MUZ'},
 {'KODU': '54', 'ADI': 'MNV.NANE'},
 {'KODU': '09397', 'ADI': 'MNV.NAR BEYAZ'},
 {'KODU': '01302', 'ADI': 'MNV.NAR KIRMIZI'},
 {'KODU': '052827', 'ADI': 'MNV.NEKTARI'},
 {'KODU': '056065', 'ADI': 'MNV.PEMBE DOMATES'},
 {'KODU': '2901187', 'ADI': 'MNV.PEPINO'},
 {'KODU': '09975', 'ADI': 'MNV.PIRASA'},
 {'KODU': '07602', 'ADI': 'MNV.PORTAKAL SIKMALIK'},
 {'KODU': 'B.2901021', 'ADI': 'MNV.RED GLOBE UZUM'},
 {'KODU': '05983', 'ADI': 'MNV.REYHAN'},
 {'KODU': '00012256', 'ADI': 'MNV.ROKA'},
 {'KODU': '09915', 'ADI': 'MNV.SALKIM DOMATES'},
 {'KODU': '09399', 'ADI': 'MNV.SANTA MARIA ARMUT'},
 {'KODU': '07604', 'ADI': 'MNV.SARI HAVUC'},
 {'KODU': '054069', 'ADI': 'MNV.SARI KAVUN'},
 {'KODU': '053793', 'ADI': 'MNV.SEKERPARE'},
 {'KODU': '98', 'ADI': 'MNV.SEMIZ OTU'},
 {'KODU': 'B.2901018', 'ADI': 'MNV.SIVRI BIBER'},
 {'KODU': '07705', 'ADI': 'MNV.SIYAH DOMATES'},
 {'KODU': '052735', 'ADI': 'MNV.SIYAH HAVUC'},
 {'KODU': 'MNV.017485', 'ADI': 'MNV.SIYAH UZUM'},
 {'KODU': '017185', 'ADI': 'MNV.SUS BIBERI'},
 {'KODU': '2901098', 'ADI': 'MNV.TATLI  PATATES'},
 {'KODU': '2901117', 'ADI': 'MNV.TAZE PATATES'},
 {'KODU': '64', 'ADI': 'MNV.TERE'},
 {'KODU': '015575', 'ADI': 'MNV.TOPAK PATLICAN'},
 {'KODU': '052113', 'ADI': 'MNV.VISNE'},
 {'KODU': '014784', 'ADI': 'MNV.WASHINGTON PORTAKAL'},
 {'KODU': '08384', 'ADI': 'MNV.YABAN MERSINI PAKET'},
 {'KODU': 'B.2901084', 'ADI': 'MNV.YAPRAK MANTAR KG'},
 {'KODU': '05694', 'ADI': 'MNV.YAYLA DOMATES'},
 {'KODU': '2909808', 'ADI': 'MNV.YAYLA ELMASI'},
 {'KODU': '05700', 'ADI': 'MNV.YENI DUNYA'},
 {'KODU': '04121', 'ADI': 'MNV.YER FISTIGI KABUKLU TAZE'},
 {'KODU': '09913', 'ADI': 'MNV.YERLI DOMATES'},
 {'KODU': '052128', 'ADI': 'MNV.YERLI SALATALIK'},
 {'KODU': '016870', 'ADI': 'MNV.YESIL KAPYA BIBER'},
 {'KODU': '38', 'ADI': 'MNV.YESIL SARIMSAK'},
 {'KODU': '053742', 'ADI': 'MNV.YESIL SILI BIBER'},
 {'KODU': '13', 'ADI': 'MNV.YESIL SOGAN'},
 {'KODU': '051277', 'ADI': 'MNV.ZENCEFIL'},
 {'KODU': '09937', 'ADI': 'MNV.INCIR'}]

if "site_giris_yapildi" not in st.session_state:
    st.session_state.site_giris_yapildi = False
if "aktif_rol" not in st.session_state:
    st.session_state.aktif_rol = "🏬 Şube Girişi"
if "giris_yapilan_sube" not in st.session_state:
    st.session_state.giris_yapilan_sube = None
if "hal_authed" not in st.session_state:
    st.session_state.hal_authed = False
if "admin_authed" not in st.session_state:
    st.session_state.admin_authed = False


def simdi_tr():
    """Sunucu hangi ülkede olursa olsun Türkiye saatini döndürür."""
    return datetime.now(ISTANBUL_TZ)



def kayit_ozeti(kayitlar):
    """Kayıtları veri tipi farklarından etkilenmeden karşılaştırmak için kararlı biçimde özetler."""
    sayisal_alanlar = {"dağıtılan_miktar", "siparis_miktari"}
    metin_alanlari = {"sube", "tarih", "urun_kodu", "urun_adi", "mevcut_stok"}

    normalize = []
    for kayit in kayitlar or []:
        temiz = {}
        for alan in sorted(kayit.keys()):
            deger = kayit.get(alan)
            if alan in sayisal_alanlar:
                try:
                    # Supabase 1, 1.0 veya "1.00" döndürse de aynı kabul edilir.
                    temiz[alan] = f"{float(deger or 0):.6f}"
                except (TypeError, ValueError):
                    temiz[alan] = "0.000000"
            elif alan in metin_alanlari:
                temiz[alan] = "" if deger is None else str(deger).strip()
            else:
                temiz[alan] = deger
        normalize.append(temiz)

    normalize.sort(key=lambda x: json.dumps(x, ensure_ascii=False, sort_keys=True, default=str))
    payload = json.dumps(normalize, ensure_ascii=False, sort_keys=True, default=str, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def baglanti_kontrolu():
    """Supabase bağlantısını hafif bir sorguyla kontrol eder."""
    try:
        supabase.table("siparisler").select("urun_kodu").limit(1).execute()
        return True
    except Exception:
        return False


def hal_taslagini_guncelle(tarih, urun_kodu):
    """Hal panelindeki alanları ürünler arasında geçişte kaybolmaması için kalıcı oturum taslağına aktarır."""
    taslaklar = st.session_state.setdefault("hal_taslaklari", {})
    taslak_anahtari = f"{tarih}|{urun_kodu}"
    toplam_key = f"hal_toplam_{tarih}_{urun_kodu}"
    dagitim = {}
    for sube in SUBE_LISTESI:
        widget_key = f"hal_dag_{tarih}_{sube}_{urun_kodu}"
        dagitim[sube] = float(st.session_state.get(widget_key, 0.0) or 0.0)
    taslaklar[taslak_anahtari] = {
        "hal_toplam": float(st.session_state.get(toplam_key, 0.0) or 0.0),
        "dagitim": dagitim,
    }


def guvenli_sorgu(islem_adi, fn):
    """Supabase işlemlerini kullanıcı dostu hata yönetimiyle çalıştırır."""
    try:
        return fn()
    except RuntimeError as exc:
        mesaj = str(exc)
        if mesaj.startswith("ÇAKIŞMA:"):
            st.warning(f"⚠️ {mesaj.replace('ÇAKIŞMA:', '').strip()}")
        else:
            st.error(f"❌ {islem_adi} sırasında bir hata oluştu: {mesaj}")
        return None
    except Exception as exc:
        hata_detayi = str(exc).strip() or exc.__class__.__name__
        st.error(f"❌ {islem_adi} sırasında hata oluştu: {hata_detayi}")
        st.caption("ℹ️ Bağlantı aktif olsa bile veritabanı yetkisi, tablo kuralı veya kayıt biçimi nedeniyle bu hata oluşabilir.")
        return None


def guvenli_veri_oku(islem_adi, fn, varsayilan=None):
    """Okuma sorgusu başarısız olursa sayfanın tamamen çökmesini önler."""
    try:
        return fn()
    except Exception:
        st.error(f"❌ {islem_adi} sırasında veri alınamadı. İnternet bağlantınızı kontrol edip tekrar deneyin.")
        return [] if varsayilan is None else varsayilan




def islem_logu_yaz(kullanici, rol, sube, islem, tablo, tarih, urun_kodu=None, urun_adi=None, eski_deger=None, yeni_deger=None, detay=None):
    """İşlem geçmişini kaydeder. Log tablosu sorunu ana işlemi durdurmaz."""
    log_zamani = simdi_tr().isoformat()
    kayit = {
        # Supabase tablosunda geçmişten kalan tarih_saat alanı da NOT NULL olduğu için
        # her iki zaman alanına aynı Türkiye saati gönderilir.
        "tarih_saat": log_zamani,
        "islem_zamani": log_zamani,
        "kullanici": str(kullanici or "Bilinmiyor"),
        "rol": str(rol or "Bilinmiyor"),
        "sube": str(sube or ""),
        "islem": str(islem),
        "tablo": str(tablo),
        "kayit_tarihi": str(tarih),
        "urun_kodu": str(urun_kodu or ""),
        "urun_adi": str(urun_adi or ""),
        "eski_deger": json.dumps(eski_deger, ensure_ascii=False, default=str) if eski_deger is not None else None,
        "yeni_deger": json.dumps(yeni_deger, ensure_ascii=False, default=str) if yeni_deger is not None else None,
        "detay": str(detay or ""),
    }
    try:
        supabase.table("islem_loglari").insert(kayit).execute()
        return True
    except Exception:
        return False


def kayitlari_kodla(kayitlar, deger_alanlari):
    """Log karşılaştırması için kayıtları ürün+şube bazında sözlüğe dönüştürür."""
    sonuc = {}
    for r in kayitlar or []:
        anahtar = (str(r.get("urun_kodu", "")), str(r.get("sube", "")))
        sonuc[anahtar] = {alan: r.get(alan) for alan in deger_alanlari}
        sonuc[anahtar]["urun_adi"] = r.get("urun_adi", "")
    return sonuc


def degisiklik_loglarini_yaz(eski, yeni, kullanici, rol, varsayilan_sube, tablo, tarih, deger_alanlari, islem_adi):
    """Eklenen, güncellenen ve silinen satırları ayrı ayrı loglar."""
    eski_map = kayitlari_kodla(eski, deger_alanlari)
    yeni_map = kayitlari_kodla(yeni, deger_alanlari)
    for anahtar in sorted(set(eski_map) | set(yeni_map)):
        once = eski_map.get(anahtar)
        sonra = yeni_map.get(anahtar)
        if once == sonra:
            continue
        kod, kayit_sube = anahtar
        urun_adi = (sonra or once or {}).get("urun_adi", "")
        if once is None:
            hareket = f"{islem_adi} - Eklendi"
        elif sonra is None:
            hareket = f"{islem_adi} - Silindi"
        else:
            hareket = f"{islem_adi} - Güncellendi"
        islem_logu_yaz(
            kullanici=kullanici,
            rol=rol,
            sube=kayit_sube or varsayilan_sube,
            islem=hareket,
            tablo=tablo,
            tarih=tarih,
            urun_kodu=kod,
            urun_adi=urun_adi,
            eski_deger=once,
            yeni_deger=sonra,
        )


def siparis_oturumunu_temizle(sube, tarih):
    """Sipariş silindikten sonra eski taslak ve widget değerlerini oturumdan kaldırır."""
    onekler = (
        f"siparis_taslak_{sube}_{tarih}",
        f"siparis_snapshot_{sube}_{tarih}",
        f"urun_arama_{sube}_{tarih}",
        f"dolu_{sube}_{tarih}_",
        f"stok_{sube}_{tarih}_",
        f"sip_{sube}_{tarih}_",
        f"urun_not_{sube}_{tarih}_",
        f"genel_siparis_notu_{sube}_{tarih}",
        f"iptal_onay_{sube}",
    )
    for anahtar in list(st.session_state.keys()):
        if any(anahtar == onek or anahtar.startswith(onek) for onek in onekler):
            del st.session_state[anahtar]


def tum_oturumlari_kapat():
    st.session_state.site_giris_yapildi = False
    st.session_state.giris_yapilan_sube = None
    st.session_state.hal_authed = False
    st.session_state.admin_authed = False


def hal_dagitimini_degistir(tarih, urun_kodu, yeni_kayitlar, beklenen_ozet=None, kullanici="Hal Yetkilisi"):

    """Aynı tarih/ürün için mükerrer kayıt ve eşzamanlı veri ezilmesini önler."""
    try:
        eski = supabase.table("hal_dagitim").select(
            "sube,tarih,urun_kodu,urun_adi,dağıtılan_miktar"
        ).eq("tarih", tarih).eq("urun_kodu", urun_kodu).execute().data or []
    except Exception as exc:
        raise RuntimeError(f"Mevcut dağıtım kaydı okunamadı: {exc}") from exc

    # Hal panelinde Streamlit her alan değişiminde sayfayı yeniden çalıştırdığı için
    # oturumdaki eski özet, aynı kullanıcının bir önceki kaydıyla zaman zaman
    # farklı kalabiliyor ve yanlış çakışma uyarısı üretebiliyordu. Hal dağıtımı
    # yetkili panelden yönetildiğinden güncel veriyi esas alıp kayda devam ediyoruz.
    # Sipariş ekranındaki gerçek eşzamanlılık kontrolü ise korunmaktadır.
    if beklenen_ozet is not None and kayit_ozeti(eski) != beklenen_ozet:
        beklenen_ozet = kayit_ozeti(eski)

    try:
        supabase.table("hal_dagitim").delete().eq("tarih", tarih).eq(
            "urun_kodu", urun_kodu
        ).execute()
    except Exception as exc:
        raise RuntimeError(f"Eski dağıtım kaydı silinemedi: {exc}") from exc

    try:
        if yeni_kayitlar:
            supabase.table("hal_dagitim").insert(yeni_kayitlar).execute()
        degisiklik_loglarini_yaz(
            eski, yeni_kayitlar, kullanici, "Hal", "", "hal_dagitim", tarih,
            ["dağıtılan_miktar"], "Hal dağıtımı"
        )
        return True
    except Exception as exc:
        # Yeni kayıt başarısız olursa eski kayıtları geri yüklemeyi dene.
        geri_yukleme_hatasi = None
        try:
            supabase.table("hal_dagitim").delete().eq("tarih", tarih).eq(
                "urun_kodu", urun_kodu
            ).execute()
            if eski:
                supabase.table("hal_dagitim").insert(eski).execute()
        except Exception as rollback_exc:
            geri_yukleme_hatasi = rollback_exc

        mesaj = f"Yeni dağıtım kaydı eklenemedi: {exc}"
        if geri_yukleme_hatasi is not None:
            mesaj += f" | Eski kayıt da geri yüklenemedi: {geri_yukleme_hatasi}"
        raise RuntimeError(mesaj) from exc


def sube_siparisini_degistir(sube, tarih, yeni_kayitlar, beklenen_ozet=None, kullanici=None):

    """Şube siparişini günceller; çakışma veya hata halinde veri kaybını engeller."""
    eski = supabase.table("siparisler").select("sube,tarih,urun_kodu,urun_adi,mevcut_stok,siparis_miktari").eq("sube", sube).eq("tarih", tarih).execute().data or []
    if beklenen_ozet is not None and kayit_ozeti(eski) != beklenen_ozet:
        raise RuntimeError("ÇAKIŞMA: Bu sipariş başka bir kullanıcı tarafından değiştirildi. Sayfayı yenileyip güncel veriyi kontrol edin.")
    try:
        supabase.table("siparisler").delete().eq("sube", sube).eq("tarih", tarih).execute()
        if yeni_kayitlar:
            supabase.table("siparisler").insert(yeni_kayitlar).execute()
        degisiklik_loglarini_yaz(
            eski, yeni_kayitlar, kullanici or sube, "Şube", sube, "siparisler", tarih,
            ["mevcut_stok", "siparis_miktari"], "Şube siparişi"
        )
        return True
    except Exception:
        try:
            supabase.table("siparisler").delete().eq("sube", sube).eq("tarih", tarih).execute()
            if eski:
                supabase.table("siparisler").insert(eski).execute()
        finally:
            raise


def generate_hal_excel(urun_adi, urun_kodu, hal_toplam, dagitim_dict, kalan, tarih_str):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hal_Dagitim_Listesi"
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_title = Font(name="Calibri", size=14, bold=True)
    font_normal = Font(name="Calibri", size=10)

    ws.cell(row=1, column=1, value="YALÇIN MARKETLER ZİNCİRİ - HAL MALI ŞUBE DAĞITIM LİSTESİ").font = font_title
    ws.cell(row=2, column=1, value=f"Tarih: {tarih_str} | Ürün: {urun_adi} (Kod: {urun_kodu}) | Halden Alınan: {hal_toplam:.0f} Kasa").font = font_bold

    headers = ["Şube Adı", "Dağıtılan Miktar (Kasa)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = font_bold
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 5
    toplam_dagitilan = 0
    for sube, miktar in dagitim_dict.items():
        if miktar > 0:
            c1 = ws.cell(row=row_idx, column=1, value=sube)
            c2 = ws.cell(row=row_idx, column=2, value=f"{miktar:.0f} Kasa")
            c1.font = font_normal
            c2.font = font_normal
            c2.alignment = Alignment(horizontal="center")
            c1.border = border
            c2.border = border
            toplam_dagitilan += miktar
            row_idx += 1

    ws.cell(row=row_idx, column=1, value="TOPLAM DAĞITILAN").font = font_bold
    c_tot = ws.cell(row=row_idx, column=2, value=f"{toplam_dagitilan:.0f} Kasa")
    c_tot.font = font_bold
    c_tot.alignment = Alignment(horizontal="center")
    ws.cell(row=row_idx, column=1).border = border
    c_tot.border = border

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    wb.save(output)
    return output.getvalue()


def generate_toplu_hal_excel(tarih_sorgu_str):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Sevkiyat_Matris"
    ws1.page_setup.orientation = ws1.ORIENTATION_LANDSCAPE
    ws1.page_setup.paperSize = ws1.PAPERSIZE_A4
    ws1.sheet_properties.pageSetUpPr.fitToPage = True
    ws1.page_setup.fitToWidth = 1
    ws1.page_setup.fitToHeight = 0
    
    try:
        res = supabase.table("hal_dagitim").select("sube, urun_kodu, urun_adi, dağıtılan_miktar").eq("tarih", tarih_sorgu_str).execute()
    except Exception:
        return None
    if not res.data:
        return None

    df_hal = pd.DataFrame(res.data)
    df_hal['dağıtılan_miktar'] = pd.to_numeric(df_hal['dağıtılan_miktar'], errors='coerce').fillna(0)
    df_hal = df_hal[df_hal['dağıtılan_miktar'] > 0]
    if df_hal.empty:
        return None

    thin = Side(border_style="thin", color="D3D3D3")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_title = Font(name="Calibri", size=13, bold=True)
    font_normal = Font(name="Calibri", size=9)

    ws1.cell(row=1, column=1, value=f"YALÇIN MARKETLER ZİNCİRİ - SEVKİYAT DAĞITIM MATRİSİ ({tarih_sorgu_str})").font = font_title

    pivot_hal = pd.pivot_table(
        df_hal, values='dağıtılan_miktar', index=['urun_kodu', 'urun_adi'], columns=['sube'], aggfunc='sum', fill_value=0
    )
    pivot_hal['TOPLAM SEVK'] = pivot_hal.sum(axis=1)

    ws1.cell(row=3, column=1, value="Ürün Kodu").font = font_bold
    ws1.cell(row=3, column=2, value="Ürün Adı").font = font_bold
    ws1.cell(row=3, column=1).fill = header_fill
    ws1.cell(row=3, column=2).fill = header_fill
    ws1.cell(row=3, column=1).border = border
    ws1.cell(row=3, column=2).border = border

    col_idx = 3
    sube_cols = [c for c in pivot_hal.columns if c != 'TOPLAM SEVK'] + ['TOPLAM SEVK']
    for sube in sube_cols:
        cell = ws1.cell(row=3, column=col_idx, value=sube)
        cell.font = font_bold
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_idx += 1

    row_idx = 4
    for (kodu, adi), r_data in pivot_hal.iterrows():
        ws1.cell(row=row_idx, column=1, value=str(kodu)).font = font_normal
        ws1.cell(row=row_idx, column=2, value=str(adi)).font = font_normal
        ws1.cell(row=row_idx, column=1).border = border
        ws1.cell(row=row_idx, column=2).border = border

        c_idx = 3
        for sube in sube_cols:
            val = r_data[sube]
            val_str = f"{int(val)} Kasa" if val > 0 else ""
            c = ws1.cell(row=row_idx, column=c_idx, value=val_str)
            c.font = font_bold if sube == 'TOPLAM SEVK' else font_normal
            c.alignment = Alignment(horizontal="center")
            c.border = border
            c_idx += 1
        row_idx += 1

    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 25
    for c in range(3, col_idx):
        ws1.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 13

    wb.save(output)
    return output.getvalue()


# ŞUBE SİPARİŞ MATRİSİNİ EXCEL'E AKTARMA FONKSİYONU
# format_tipi: "standart", "a3" veya "a4_bolunmus"
def generate_sube_siparis_excel(tarih_sorgu_str, df_wide, format_tipi="standart"):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    thin = Side(border_style="thin", color="B7C9B7")
    medium = Side(border_style="medium", color="5B7F5B")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    strong_border = Border(top=medium, left=medium, right=medium, bottom=medium)
    title_fill = PatternFill(start_color="1F6B3A", end_color="1F6B3A", fill_type="solid")
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    subheader_fill = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    zebra_fill = PatternFill(start_color="F7FBF7", end_color="F7FBF7", fill_type="solid")
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Calibri", size=11, bold=True, color="1F1F1F")
    font_bold = Font(name="Calibri", size=10, bold=True)
    font_normal = Font(name="Calibri", size=9)
    font_print = Font(name="Calibri", size=10)

    try:
        rapor_tarihi = datetime.strptime(tarih_sorgu_str, "%Y-%m-%d").strftime("%d.%m.%Y")
    except ValueError:
        rapor_tarihi = tarih_sorgu_str

    olusturma_zamani = simdi_tr().strftime("%d.%m.%Y %H:%M")
    toplam_urun = len(df_wide)
    toplam_siparis = float(pd.to_numeric(df_wide.get("toplam_sip", 0), errors="coerce").fillna(0).sum())

    def sayisal_siparis(deger):
        try:
            return float(deger) if deger not in ("", "-", None) else 0.0
        except (TypeError, ValueError):
            return 0.0

    def sayfa_olustur(sheet_name, subeler, genel_toplam_ekle=True, kagit="A4", tek_sayfaya_sigdir=False):
        ws = wb.create_sheet(title=sheet_name[:31])
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "C5"
        ws.auto_filter.ref = None
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A3 if kagit == "A3" else ws.PAPERSIZE_A4
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.2
        ws.page_margins.right = 0.2
        ws.page_margins.top = 0.45
        ws.page_margins.bottom = 0.45
        ws.page_margins.header = 0.15
        ws.page_margins.footer = 0.2
        ws.print_title_rows = "3:4"
        ws.oddFooter.center.text = "Yalçın Marketler Zinciri - Manav Sipariş ve Stok Yönetim Portalı"
        ws.oddFooter.right.text = "Sayfa &P / &N"
        ws.oddFooter.left.text = f"Rapor: {rapor_tarihi}"
        ws.sheet_properties.pageSetUpPr.autoPageBreaks = False
        ws.page_setup.scale = 80 if tek_sayfaya_sigdir else None

        son_kolon = 2 + (len(subeler) * 2) + (2 if genel_toplam_ekle else 0)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=son_kolon)
        t = ws.cell(row=1, column=1, value="YALÇIN MARKETLER ZİNCİRİ - MANAV STOK VE SİPARİŞ RAPORU")
        t.font = font_title
        t.fill = title_fill
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=son_kolon)
        info = ws.cell(
            row=2,
            column=1,
            value=(f"Rapor Tarihi: {rapor_tarihi}   |   Oluşturma: {olusturma_zamani}   |   "
                   f"Toplam Ürün: {toplam_urun}   |   Toplam Sipariş: {toplam_siparis:.0f} Kasa")
        )
        info.font = font_subtitle
        info.alignment = Alignment(horizontal="center", vertical="center")
        info.fill = subheader_fill
        ws.row_dimensions[2].height = 22

        for col, value in ((1, "Ürün Kodu"), (2, "Ürün Adı")):
            c = ws.cell(row=3, column=col, value=value)
            c.font = font_bold
            c.fill = header_fill
            c.border = strong_border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)

        c_idx = 3
        for s_name in subeler:
            ws.merge_cells(start_row=3, start_column=c_idx, end_row=3, end_column=c_idx + 1)
            c = ws.cell(row=3, column=c_idx, value=s_name)
            c.font = font_bold
            c.fill = header_fill
            c.border = strong_border
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.cell(row=3, column=c_idx + 1).border = strong_border

            for offset, text in ((0, "Stok"), (1, "Sip.")):
                sc = ws.cell(row=4, column=c_idx + offset, value=text)
                sc.font = font_bold
                sc.fill = subheader_fill
                sc.border = border
                sc.alignment = Alignment(horizontal="center", vertical="center")
            c_idx += 2

        if genel_toplam_ekle:
            ws.merge_cells(start_row=3, start_column=c_idx, end_row=3, end_column=c_idx + 1)
            c = ws.cell(row=3, column=c_idx, value="GENEL TOPLAM")
            c.font = font_bold
            c.fill = total_fill
            c.border = strong_border
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=3, column=c_idx + 1).border = strong_border
            for offset, text in ((0, "Top. Stok / RD"), (1, "Top. Sipariş")):
                sc = ws.cell(row=4, column=c_idx + offset, value=text)
                sc.font = font_bold
                sc.fill = total_fill
                sc.border = border
                sc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        row_idx = 5
        for sira, (_, r) in enumerate(df_wide.iterrows(), start=1):
            row_fill = zebra_fill if sira % 2 == 0 else None
            ws.cell(row=row_idx, column=1, value=str(r["urun_kodu"]))
            ws.cell(row=row_idx, column=2, value=str(r["urun_adi"]))
            for col in (1, 2):
                c = ws.cell(row=row_idx, column=col)
                c.font = font_print if format_tipi != "standart" else font_normal
                c.border = border
                c.alignment = Alignment(vertical="center", wrap_text=(col == 2))
                if row_fill:
                    c.fill = row_fill

            curr_c = 3
            for s_name in subeler:
                stok_val = r.get(f"{s_name}_stok", "-")
                sip_val = r.get(f"{s_name}_sip", "-")
                values = (str(stok_val), sayisal_siparis(sip_val) or "")
                for offset, value in enumerate(values):
                    c = ws.cell(row=row_idx, column=curr_c + offset, value=value)
                    c.font = font_print if format_tipi != "standart" else font_normal
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    c.border = border
                    if row_fill:
                        c.fill = row_fill
                curr_c += 2

            if genel_toplam_ekle:
                gt_stok = r.get("toplam_stok", "-")
                gt_sip = sayisal_siparis(r.get("toplam_sip", 0))
                for offset, value in ((0, str(gt_stok)), (1, gt_sip or "")):
                    c = ws.cell(row=row_idx, column=curr_c + offset, value=value)
                    c.font = font_bold
                    c.fill = total_fill
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    c.border = border
            ws.row_dimensions[row_idx].height = 19 if format_tipi == "standart" else 22
            row_idx += 1

        # Şube toplamları özeti
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="ŞUBE SİPARİŞ ÖZETİ").font = font_bold
        ws.cell(row=row_idx, column=1).fill = title_fill
        ws.cell(row=row_idx, column=1).font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
        row_idx += 1

        ozet_baslangic = row_idx
        for s_name in subeler:
            toplam = sum(sayisal_siparis(v) for v in df_wide.get(f"{s_name}_sip", []))
            ws.cell(row=row_idx, column=1, value=s_name)
            ws.cell(row=row_idx, column=2, value=toplam)
            ws.cell(row=row_idx, column=2).number_format = '0 "Kasa"'
            for col in (1, 2):
                c = ws.cell(row=row_idx, column=col)
                c.border = border
                c.font = font_bold if col == 2 else font_normal
                c.alignment = Alignment(horizontal="center" if col == 2 else "left")
            row_idx += 1
        ws.cell(row=row_idx, column=1, value="GENEL TOPLAM")
        ws.cell(row=row_idx, column=2, value=toplam_siparis)
        ws.cell(row=row_idx, column=2).number_format = '0 "Kasa"'
        for col in (1, 2):
            c = ws.cell(row=row_idx, column=col)
            c.font = font_bold
            c.fill = total_fill
            c.border = strong_border

        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 30 if format_tipi != "a4_bolunmus" else 34
        for col in range(3, son_kolon + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 10 if format_tipi == "a4_bolunmus" else 9

        ws.auto_filter.ref = f"A4:{openpyxl.utils.get_column_letter(son_kolon)}{row_idx - len(subeler) - 3}"
        ws.print_area = f"A1:{openpyxl.utils.get_column_letter(son_kolon)}{row_idx}"
        ws.sheet_properties.outlinePr.summaryBelow = True
        ws.sheet_view.zoomScale = 85
        return ws

    if format_tipi == "a4_bolunmus":
        ilk_grup = SUBE_LISTESI[:5]
        ikinci_grup = SUBE_LISTESI[5:]
        sayfa_olustur("A4_1_Ilk_5_Sube", ilk_grup, genel_toplam_ekle=False, kagit="A4")
        sayfa_olustur("A4_2_Diger_Subeler", ikinci_grup, genel_toplam_ekle=True, kagit="A4")
    elif format_tipi == "a3":
        sayfa_olustur("A3_Tek_Sayfa", SUBE_LISTESI, genel_toplam_ekle=True, kagit="A3", tek_sayfaya_sigdir=True)
    else:
        sayfa_olustur("Duzenlenebilir_Matris", SUBE_LISTESI, genel_toplam_ekle=True, kagit="A4")

    wb.save(output)
    output.seek(0)
    return output.getvalue()



# A4 DİKEY - EN FAZLA 2 SAYFA - KOMPAKT STOK/SİPARİŞ RAPORU
# Şube sırası, onaylanan örnek dosyadaki sırayla sabittir.
def generate_sube_siparis_dikey_2_sayfa(tarih_sorgu_str, df_wide):
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dikey_2_Sayfa"
    ws.sheet_view.showGridLines = False

    # Kullanıcının onayladığı çıktıdaki şube sırası.
    sube_sirasi = [
        ("Aşiyan", "Aşiyan"),
        ("Metin Tamer", "Metin T."),
        ("Hacı Osmanlı", "Hacı O."),
        ("Salı Yolu", "Salı Y."),
        ("Bulvar", "Bulvar"),
        ("Düziçi Çarşı", "Düziçi"),
        ("Kadiri Yolu", "Kadirli"),
        ("Zeytinlik", "Zeytinlik"),
        ("Raufbey", "Rauf."),
        ("Eyup Sultan", "Eyüp S."),
        ("Nahır Yolu", "Nahır Y."),
    ]

    ince = Side(border_style="thin", color="C8C8C8")
    kenarlik = Border(top=ince, left=ince, right=ince, bottom=ince)
    baslik_dolgu = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    toplam_dolgu = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    zebra_dolgu = PatternFill(start_color="F7FBF7", end_color="F7FBF7", fill_type="solid")
    baslik_font = Font(name="Calibri", size=7, bold=True)
    normal_font = Font(name="Calibri", size=7)
    toplam_font = Font(name="Calibri", size=7, bold=True)

    def temiz_stok(deger):
        if deger is None:
            return "-"
        metin = str(deger).strip()
        if not metin or metin == "-":
            return "-"
        if metin.lower() == "reyon dolu":
            return "RD"
        try:
            sayi = float(metin)
            return str(int(sayi)) if sayi.is_integer() else f"{sayi:g}"
        except (TypeError, ValueError):
            return metin

    def temiz_siparis(deger):
        if deger is None or str(deger).strip() in ("", "-"):
            return "-"
        try:
            sayi = float(deger)
            if sayi == 0:
                return "-"
            return str(int(sayi)) if sayi.is_integer() else f"{sayi:g}"
        except (TypeError, ValueError):
            return str(deger).strip()

    def kompakt_toplam_stok(deger):
        metin = str(deger or "-").strip()
        metin = metin.replace(" Kasa", "")
        metin = metin.replace(" (+", "+").replace(" RD)", "RD")
        return metin or "-"

    headers = ["Ürün Adı"] + [kisa for _, kisa in sube_sirasi] + ["Top. Stok/RD", "Top. Sip."]
    ws.append(headers)

    for col_idx, baslik in enumerate(headers, start=1):
        hucre = ws.cell(row=1, column=col_idx, value=baslik)
        hucre.font = baslik_font
        hucre.fill = toplam_dolgu if col_idx >= 13 else baslik_dolgu
        hucre.border = kenarlik
        hucre.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 24

    for sira, (_, row) in enumerate(df_wide.iterrows(), start=1):
        satir = [str(row.get("urun_adi", ""))]
        for sube_adi, _ in sube_sirasi:
            stok = temiz_stok(row.get(f"{sube_adi}_stok", "-"))
            siparis = temiz_siparis(row.get(f"{sube_adi}_sip", "-"))
            satir.append(f"{stok}/{siparis}")
        satir.extend([
            kompakt_toplam_stok(row.get("toplam_stok", "-")),
            temiz_siparis(row.get("toplam_sip", 0)),
        ])
        ws.append(satir)
        excel_satir = ws.max_row
        for col_idx in range(1, 15):
            hucre = ws.cell(row=excel_satir, column=col_idx)
            hucre.font = toplam_font if col_idx >= 13 else normal_font
            hucre.fill = toplam_dolgu if col_idx >= 13 else (zebra_dolgu if sira % 2 == 0 else PatternFill(fill_type=None))
            hucre.border = kenarlik
            hucre.alignment = Alignment(
                horizontal="left" if col_idx == 1 else "center",
                vertical="center",
                wrap_text=(col_idx == 1),
            )
        ws.row_dimensions[excel_satir].height = 15

    # Örnek dosyadaki dar ve okunabilir sütun düzeni.
    ws.column_dimensions["A"].width = 22
    for col in range(2, 13):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 5.6
    ws.column_dimensions["M"].width = 8.5
    ws.column_dimensions["N"].width = 7.5

    # A4 dikey, genişlik tek sayfa; uzunluk en fazla iki sayfa.
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 2
    ws.page_margins.left = 0.15
    ws.page_margins.right = 0.15
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.10
    ws.page_margins.footer = 0.10
    ws.print_title_rows = "1:1"
    ws.print_area = f"A1:N{ws.max_row}"
    ws.oddFooter.left.text = f"Tarih: {tarih_sorgu_str}"
    ws.oddFooter.center.text = "RD = Reyon Dolu | Hücre: Stok/Sipariş"
    ws.oddFooter.right.text = "Sayfa &P / &N"
    ws.sheet_view.zoomScale = 80

    # Ürünleri mümkün olduğunca eşit biçimde iki sayfaya ayır.
    if ws.max_row > 44:
        orta_satir = 1 + ((ws.max_row - 1 + 1) // 2)
        ws.row_breaks.append(openpyxl.worksheet.pagebreak.Break(id=orta_satir))

    wb.save(output)
    output.seek(0)
    return output.getvalue()


def siparis_notlarini_oku(sube, baslangic_tarihi, bitis_tarihi=None):
    """Şube sipariş notlarını güvenli şekilde okur."""
    bitis_tarihi = bitis_tarihi or baslangic_tarihi
    return guvenli_veri_oku(
        "Sipariş notlarını okuma",
        lambda: supabase.table("siparis_notlari")
            .select("sube,tarih,urun_kodu,urun_notu,genel_not")
            .eq("sube", sube)
            .gte("tarih", str(baslangic_tarihi))
            .lte("tarih", str(bitis_tarihi))
            .execute().data or [],
        varsayilan=[],
    )


def siparis_notlarini_kaydet(sube, tarih, urun_notlari, genel_not):
    """Ürün ve genel sipariş notlarını tarih/şube bazında yeniler."""
    mevcut = supabase.table("siparis_notlari").select("*").eq("sube", sube).eq("tarih", tarih).execute().data or []
    try:
        supabase.table("siparis_notlari").delete().eq("sube", sube).eq("tarih", tarih).execute()
        yeni = []
        temiz_genel_not = str(genel_not or "").strip()
        for urun_kodu, urun_notu in (urun_notlari or {}).items():
            temiz_not = str(urun_notu or "").strip()
            if temiz_not:
                yeni.append({
                    "sube": sube,
                    "tarih": tarih,
                    "urun_kodu": str(urun_kodu),
                    "urun_notu": temiz_not,
                    "genel_not": temiz_genel_not,
                })
        # Ürün notu yoksa genel notu özel bir satırda sakla.
        if temiz_genel_not and not yeni:
            yeni.append({
                "sube": sube,
                "tarih": tarih,
                "urun_kodu": "__GENEL__",
                "urun_notu": "",
                "genel_not": temiz_genel_not,
            })
        if yeni:
            supabase.table("siparis_notlari").insert(yeni).execute()
        return True
    except Exception:
        try:
            supabase.table("siparis_notlari").delete().eq("sube", sube).eq("tarih", tarih).execute()
            if mevcut:
                supabase.table("siparis_notlari").insert(mevcut).execute()
        finally:
            raise


def notlari_kayitlara_ekle(kayitlar, not_kayitlari):
    """Not kayıtlarını sipariş satırlarına birleştirir ve genel notu döndürür."""
    not_map = {}
    genel_not = ""
    for n in not_kayitlari or []:
        kod = str(n.get("urun_kodu") or "")
        if kod and kod != "__GENEL__":
            not_map[kod] = str(n.get("urun_notu") or "")
        if not genel_not and str(n.get("genel_not") or "").strip():
            genel_not = str(n.get("genel_not") or "").strip()
    birlesik = []
    for kayit in kayitlar or []:
        kod = str(kayit.get("urun_kodu") or "")
        birlesik.append({**kayit, "urun_notu": not_map.get(kod, "")})
    return birlesik, genel_not

def generate_sube_tek_siparis_excel(sube, tarih_str, kayitlar, genel_not=""):
    """Tek bir şubenin seçilen tarihli siparişini A4'e uygun Excel olarak üretir."""
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Siparis_Dokumu"
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45

    thin = Side(border_style="thin", color="B7B7B7")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    title_fill = PatternFill(start_color="1F6B3A", end_color="1F6B3A", fill_type="solid")
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")

    try:
        gorunen_tarih = datetime.strptime(tarih_str, "%Y-%m-%d").strftime("%d.%m.%Y")
    except ValueError:
        gorunen_tarih = tarih_str

    temiz = []
    for r in kayitlar or []:
        try:
            miktar = float(r.get("siparis_miktari") or 0)
        except (TypeError, ValueError):
            miktar = 0.0
        if miktar > 0:
            temiz.append({**r, "siparis_miktari": miktar})

    toplam_kasa = sum(r["siparis_miktari"] for r in temiz)
    siparis_no = f"YM-{tarih_str.replace('-', '')}-{SUBE_LISTESI.index(sube)+1:02d}" if sube in SUBE_LISTESI else f"YM-{tarih_str.replace('-', '')}"

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "YALÇIN MARKETLER ZİNCİRİ - ŞUBE SİPARİŞ DÖKÜMÜ"
    c.font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    c.fill = title_fill
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    bilgi = [
        ("Şube", sube), ("Sipariş Tarihi", gorunen_tarih),
        ("Sipariş No", siparis_no), ("Oluşturma", simdi_tr().strftime("%d.%m.%Y %H:%M")),
    ]
    for i, (etiket, deger) in enumerate(bilgi, start=2):
        ws.cell(i, 1, etiket).font = Font(bold=True)
        ws.cell(i, 2, deger)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)

    headers = ["Ürün Kodu", "Ürün Adı", "Mevcut Stok", "Sipariş (Kasa)", "Ürün Notu"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(7, col, header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row = 8
    for r in sorted(temiz, key=lambda x: str(x.get("urun_adi", ""))):
        values = [r.get("urun_kodu", ""), r.get("urun_adi", ""), r.get("mevcut_stok", ""), r.get("siparis_miktari", 0), r.get("urun_notu", "")]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row, col, value)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col not in (2, 5) else "left", vertical="center", wrap_text=True)
        row += 1

    ws.cell(row, 1, "TOPLAM").font = Font(bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    ws.cell(row, 5, toplam_kasa).font = Font(bold=True)
    ws.cell(row, 5).number_format = '0.## "Kasa"'
    for col in range(1, 6):
        ws.cell(row, col).fill = total_fill
        ws.cell(row, col).border = border
    ws.cell(row, 1).alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 30
    ws.freeze_panes = "A8"
    ws.print_title_rows = "1:7"
    if str(genel_not or "").strip():
        row += 2
        ws.cell(row, 1, "GENEL SİPARİŞ NOTU").font = Font(bold=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(row, 1).fill = header_fill
        row += 1
        ws.cell(row, 1, str(genel_not).strip())
        ws.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=5)
        ws.cell(row, 1).alignment = Alignment(wrap_text=True, vertical="top")
        for rr in (row, row + 1):
            for cc in range(1, 6):
                ws.cell(rr, cc).border = border
    ws.print_area = f"A1:E{row + (1 if str(genel_not or '').strip() else 0)}"
    ws.oddFooter.center.text = "Yalçın Marketler Zinciri - Manav Sipariş Portalı"
    ws.oddFooter.right.text = "Sayfa &P / &N"
    wb.save(output)
    return output.getvalue()


def generate_sube_siparis_html(sube, tarih_str, kayitlar, genel_not=""):
    """Tarayıcı yazdırma penceresi için güvenli, A4 uyumlu HTML üretir."""
    import html
    try:
        gorunen_tarih = datetime.strptime(tarih_str, "%Y-%m-%d").strftime("%d.%m.%Y")
    except ValueError:
        gorunen_tarih = tarih_str
    satirlar = []
    toplam = 0.0
    for r in sorted(kayitlar or [], key=lambda x: str(x.get("urun_adi", ""))):
        try:
            miktar = float(r.get("siparis_miktari") or 0)
        except (TypeError, ValueError):
            miktar = 0.0
        if miktar <= 0:
            continue
        toplam += miktar
        miktar_text = f"{miktar:g}"
        satirlar.append(
            "<tr>"
            f"<td>{html.escape(str(r.get('urun_kodu', '')))}</td>"
            f"<td>{html.escape(str(r.get('urun_adi', '')))}</td>"
            f"<td>{html.escape(str(r.get('mevcut_stok', '')))}</td>"
            f"<td class='num'>{miktar_text}</td>"
            f"<td>{html.escape(str(r.get('urun_notu', '')))}</td>"
            "</tr>"
        )
    siparis_no = f"YM-{tarih_str.replace('-', '')}-{SUBE_LISTESI.index(sube)+1:02d}" if sube in SUBE_LISTESI else f"YM-{tarih_str.replace('-', '')}"
    return f"""
<!doctype html><html><head><meta charset='utf-8'>
<style>
body{{font-family:Arial,sans-serif;margin:0;color:#111}} .toolbar{{padding:10px 0}}
button{{background:#1f6b3a;color:white;border:0;border-radius:7px;padding:11px 18px;font-size:15px;font-weight:700;cursor:pointer}}
.sheet{{max-width:820px;margin:auto;padding:12px}} h2{{text-align:center;margin:4px 0 14px}}
.meta{{display:grid;grid-template-columns:145px 1fr;gap:5px 10px;margin-bottom:14px}} .meta b{{background:#eef5ee;padding:5px}}
table{{width:100%;border-collapse:collapse;font-size:12px}} th,td{{border:1px solid #777;padding:6px}} th{{background:#ddebf7}} .num{{text-align:center}}
tfoot td{{font-weight:bold;background:#fff2cc}} .general-note{{margin-top:14px;border:1px solid #777;padding:10px;background:#fffbe6;white-space:pre-wrap}} @media print{{.toolbar{{display:none}} .sheet{{padding:0}} @page{{size:A4 portrait;margin:12mm}}}}
</style></head><body><div class='sheet'><div class='toolbar'><button onclick='window.print()'>🖨️ YAZDIR / PDF OLARAK KAYDET</button></div>
<h2>YALÇIN MARKETLER ZİNCİRİ<br>ŞUBE SİPARİŞ DÖKÜMÜ</h2>
<div class='meta'><b>Şube</b><span>{html.escape(sube)}</span><b>Sipariş Tarihi</b><span>{html.escape(gorunen_tarih)}</span><b>Sipariş No</b><span>{html.escape(siparis_no)}</span></div>
<table><thead><tr><th>Ürün Kodu</th><th>Ürün Adı</th><th>Mevcut Stok</th><th>Sipariş (Kasa)</th><th>Ürün Notu</th></tr></thead><tbody>{''.join(satirlar)}</tbody>
<tfoot><tr><td colspan='4' style='text-align:right'>TOPLAM</td><td class='num'>{toplam:g} Kasa</td></tr></tfoot></table>
{f"<div class='general-note'><b>Genel Sipariş Notu</b><br>{html.escape(str(genel_not).strip())}</div>" if str(genel_not or '').strip() else ''}</div></body></html>
"""

def generate_tum_veri_yedegi():
    """Siparişler ve hal dağıtım tablolarının tamamını tek Excel dosyasında yedekler."""
    try:
        siparis_data = supabase.table("siparisler").select("*").order("tarih").execute().data or []
        hal_data = supabase.table("hal_dagitim").select("*").order("tarih").execute().data or []
    except Exception as exc:
        raise RuntimeError(f"Yedek verileri Supabase'den alınamadı: {exc}") from exc

    output = io.BytesIO()
    wb = openpyxl.Workbook()
    ws_bilgi = wb.active
    ws_bilgi.title = "Yedek Bilgisi"

    baslik_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    alt_fill = PatternFill(start_color="D9EAF7", end_color="D9EAF7", fill_type="solid")
    beyaz_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    kalin_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=10)
    thin = Side(border_style="thin", color="B7B7B7")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    yedek_zamani = simdi_tr()
    ws_bilgi.merge_cells("A1:D1")
    ws_bilgi["A1"] = "YALÇIN MARKETLER ZİNCİRİ - MANAV PORTALI VERİ YEDEĞİ"
    ws_bilgi["A1"].font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
    ws_bilgi["A1"].fill = baslik_fill
    ws_bilgi["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_bilgi.row_dimensions[1].height = 28

    bilgiler = [
        ("Yedek Tarihi", yedek_zamani.strftime("%d.%m.%Y")),
        ("Yedek Saati", yedek_zamani.strftime("%H:%M:%S")),
        ("Sipariş Kayıt Sayısı", len(siparis_data)),
        ("Hal Dağıtım Kayıt Sayısı", len(hal_data)),
        ("Toplam Kayıt", len(siparis_data) + len(hal_data)),
    ]
    for satir, (etiket, deger) in enumerate(bilgiler, start=3):
        ws_bilgi.cell(satir, 1, etiket).font = kalin_font
        ws_bilgi.cell(satir, 1).fill = alt_fill
        ws_bilgi.cell(satir, 1).border = border
        ws_bilgi.cell(satir, 2, deger).font = normal_font
        ws_bilgi.cell(satir, 2).border = border

    ws_bilgi["A10"] = "Önemli"
    ws_bilgi["A10"].font = kalin_font
    ws_bilgi["A11"] = "Bu dosyayı değiştirmeden güvenli bir klasörde saklayın. Geri yükleme gerektiğinde bu yedek kullanılabilir."
    ws_bilgi.merge_cells("A11:D12")
    ws_bilgi["A11"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_bilgi.column_dimensions["A"].width = 28
    ws_bilgi.column_dimensions["B"].width = 24
    ws_bilgi.column_dimensions["C"].width = 18
    ws_bilgi.column_dimensions["D"].width = 18

    def tablo_sayfasi_ekle(sayfa_adi, kayitlar, tercih_edilen_sutunlar):
        ws = wb.create_sheet(title=sayfa_adi)
        if kayitlar:
            mevcut_sutunlar = []
            for sutun in tercih_edilen_sutunlar:
                if any(sutun in kayit for kayit in kayitlar):
                    mevcut_sutunlar.append(sutun)
            for kayit in kayitlar:
                for sutun in kayit.keys():
                    if sutun not in mevcut_sutunlar:
                        mevcut_sutunlar.append(sutun)
        else:
            mevcut_sutunlar = tercih_edilen_sutunlar

        for col, sutun in enumerate(mevcut_sutunlar, start=1):
            hucre = ws.cell(1, col, sutun)
            hucre.font = beyaz_font
            hucre.fill = baslik_fill
            hucre.alignment = Alignment(horizontal="center", vertical="center")
            hucre.border = border

        for row, kayit in enumerate(kayitlar, start=2):
            for col, sutun in enumerate(mevcut_sutunlar, start=1):
                deger = kayit.get(sutun)
                if isinstance(deger, (dict, list)):
                    deger = json.dumps(deger, ensure_ascii=False, default=str)
                hucre = ws.cell(row, col, deger)
                hucre.font = normal_font
                hucre.border = border
                hucre.alignment = Alignment(vertical="center")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        for col, sutun in enumerate(mevcut_sutunlar, start=1):
            max_uzunluk = len(str(sutun))
            for row in range(2, min(ws.max_row, 500) + 1):
                value = ws.cell(row, col).value
                if value is not None:
                    max_uzunluk = max(max_uzunluk, len(str(value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max(max_uzunluk + 2, 12), 38)

        if not kayitlar:
            ws.cell(2, 1, "Bu tabloda henüz kayıt bulunmuyor.")

    tablo_sayfasi_ekle(
        "Siparişler",
        siparis_data,
        ["id", "sube", "tarih", "urun_kodu", "urun_adi", "mevcut_stok", "siparis_miktari", "created_at"],
    )
    tablo_sayfasi_ekle(
        "Hal Dağıtımı",
        hal_data,
        ["id", "sube", "tarih", "urun_kodu", "urun_adi", "dağıtılan_miktar", "created_at"],
    )

    wb.save(output)
    return output.getvalue(), len(siparis_data), len(hal_data), yedek_zamani


# KARŞILAMA EKRANI
if not st.session_state.site_giris_yapildi:
    st.markdown("<br>", unsafe_allow_html=True)
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        try:
            with open("logo.png", "rb") as image_file:
                encoded_string = base64.b64encode(image_file.read()).decode()
            st.markdown(f'''
                <div class="logo-card-container">
                    <img src="data:image/png;base64,{encoded_string}" class="animated-logo">
                </div>
            ''', unsafe_allow_html=True)
        except Exception:
            st.caption("ℹ️ Logo dosyası bulunamadı; sistem çalışmaya devam ediyor.")

        st.markdown('<div class="welcome-title">YALÇIN MARKETLER ZİNCİRİ</div>', unsafe_allow_html=True)
        st.markdown('<div class="welcome-sub">Manav Sipariş ve Stok Yönetim Portalı</div>', unsafe_allow_html=True)
        
        if st.button("🚀 SİSTEME GİRİŞ YAP", type="primary", use_container_width=True):
            st.session_state.site_giris_yapildi = True
            st.rerun()

else:
    st.markdown("### 📌 Sayfa Geçişi")
    m_col1, m_col2, m_col3, m_col4 = st.columns([1, 1, 1, 0.8])
    with m_col1:
        if st.button("🏬 Şube Girişi", type="primary" if st.session_state.aktif_rol == "🏬 Şube Girişi" else "secondary", use_container_width=True):
            st.session_state.aktif_rol = "🏬 Şube Girişi"
            st.rerun()
    with m_col2:
        if st.button("🚛 Hal Dağıtım Paneli", type="primary" if st.session_state.aktif_rol == "🚛 Hal Dağıtım Paneli" else "secondary", use_container_width=True):
            st.session_state.aktif_rol = "🚛 Hal Dağıtım Paneli"
            st.rerun()
    with m_col3:
        if st.button("👑 Merkez Panel", type="primary" if st.session_state.aktif_rol == "👑 Merkez Panel" else "secondary", use_container_width=True):
            st.session_state.aktif_rol = "👑 Merkez Panel"
            st.rerun()
    with m_col4:
        if st.button("🚪 Çıkış", use_container_width=True):
            tum_oturumlari_kapat()
            st.rerun()

    st.divider()
    baglanti_var = baglanti_kontrolu()
    if baglanti_var:
        st.caption("🟢 Veri bağlantısı aktif")
    else:
        st.error("🔴 Veri bağlantısı kurulamadı. Girdiğiniz alanlar bu sayfa açık kaldığı sürece korunur; bağlantı geldikten sonra tekrar kaydedin.")
    rol = st.session_state.aktif_rol

    # 1. ŞUBE SİPARİŞ GİRİŞİ
    if rol == "🏬 Şube Girişi":
        st.markdown("<h2 style='text-align: center;'>🥭 Şube Manav Sipariş Portalı</h2>", unsafe_allow_html=True)
        bugun_str = simdi_tr().strftime('%Y-%m-%d')
        st.caption(f"Tarih: {simdi_tr().strftime('%d.%m.%Y')}")

        subeler = ["-- Seçiniz --"] + SUBE_LISTESI
        secilen_sube = st.selectbox("📍 **Lütfen Şubenizi Seçin:**", subeler)

        if secilen_sube != "-- Seçiniz --":
            if st.session_state.giris_yapilan_sube != secilen_sube:
                st.info(f"🔒 **{secilen_sube}** şubesinin sipariş ekranına erişmek için lütfen şube şifrenizi giriniz.")
                s_col1, s_col2 = st.columns([2, 1])
                with s_col1:
                    girilen_pin = st.text_input(f"🔑 {secilen_sube} Şube Şifresi:", type="password", key=f"pin_input_{secilen_sube}")
                with s_col2:
                    st.markdown("<br>", unsafe_allow_html=True)
                    if st.button("Giriş Yap", type="primary", use_container_width=True):
                        if girilen_pin == SUBE_SIFRELERI.get(secilen_sube):
                            st.session_state.giris_yapilan_sube = secilen_sube
                            st.success("✅ Şifre Doğrulandı!")
                            st.rerun()
                        else:
                            st.error("❌ Hatalı Şube Şifresi!")
            else:
                st.success(f"🔓 **{secilen_sube}** Şubesi Girişi Aktif")
                if st.button("🔒 Şube Oturumunu Kapat", type="secondary"):
                    st.session_state.giris_yapilan_sube = None
                    st.rerun()

                st.divider()

                with st.expander(f"🚛 **{secilen_sube} - Halden Şubemize Ayrılan/Gelen Mal Miktarları (Bugün)**", expanded=True):
                    hal_verileri = guvenli_veri_oku(
                        "Hal dağıtım bilgilerini okuma",
                        lambda: supabase.table("hal_dagitim").select("urun_kodu, urun_adi, dağıtılan_miktar").eq("sube", secilen_sube).eq("tarih", bugun_str).execute().data or []
                    )
                    if hal_verileri:
                        hal_df = pd.DataFrame(hal_verileri)
                        hal_df['dağıtılan_miktar'] = pd.to_numeric(hal_df['dağıtılan_miktar'], errors='coerce').fillna(0)
                        hal_df = hal_df[hal_df['dağıtılan_miktar'] > 0]
                        if not hal_df.empty:
                            hal_df = hal_df.rename(columns={'urun_kodu': 'Ürün Kodu', 'urun_adi': 'Ürün Adı', 'dağıtılan_miktar': 'Gelen / Ayrılan Miktar (Kasa)'})
                            st.dataframe(hal_df, use_container_width=True, hide_index=True)
                        else:
                            st.info("ℹ️ Bugün için şubenize henüz halden yüklenen mal girişi yapılmadı.")
                    else:
                        st.info("ℹ️ Bugün için şubenize henüz halden yüklenen mal girişi yapılmadı.")

                st.divider()

                siparis_verileri = guvenli_veri_oku(
                    "Şube siparişlerini okuma",
                    lambda: supabase.table("siparisler").select("sube,tarih,urun_kodu,urun_adi,mevcut_stok,siparis_miktari").eq("sube", secilen_sube).eq("tarih", bugun_str).execute().data or []
                )
                bugun_not_kayitlari = siparis_notlarini_oku(secilen_sube, bugun_str)
                bugun_urun_notlari = {
                    str(n.get("urun_kodu")): str(n.get("urun_notu") or "")
                    for n in bugun_not_kayitlari if str(n.get("urun_kodu") or "") != "__GENEL__"
                }
                bugun_genel_not = next(
                    (str(n.get("genel_not") or "") for n in bugun_not_kayitlari if str(n.get("genel_not") or "").strip()),
                    "",
                )
                siparis_snapshot_key = f"siparis_snapshot_{secilen_sube}_{bugun_str}"
                if siparis_snapshot_key not in st.session_state:
                    st.session_state[siparis_snapshot_key] = kayit_ozeti(siparis_verileri)
                kayitli_dict = {}
                for r in siparis_verileri:
                    try:
                        sip_val = float(r['siparis_miktari']) if r['siparis_miktari'] is not None else 0.0
                    except (ValueError, TypeError):
                        sip_val = 0.0
                    kayitli_dict[r['urun_kodu']] = {
                        'stok': str(r['mevcut_stok']) if r['mevcut_stok'] is not None else "0",
                        'siparis': sip_val
                    }

                # Şube+tarih bazlı tam sipariş taslağı. Arama yalnızca görünümü filtreler;
                # kaydetme işlemi her zaman tüm ürünlerin taslağını kullanır.
                siparis_taslak_key = f"siparis_taslak_{secilen_sube}_{bugun_str}"
                if siparis_taslak_key not in st.session_state:
                    st.session_state[siparis_taslak_key] = {
                        u["KODU"]: {
                            "urun_adi": u["ADI"],
                            "stok": kayitli_dict.get(u["KODU"], {}).get("stok", "0"),
                            "siparis": float(kayitli_dict.get(u["KODU"], {}).get("siparis", 0.0)),
                            "not": bugun_urun_notlari.get(u["KODU"], ""),
                        }
                        for u in URUNLER
                    }

                siparis_taslagi = st.session_state[siparis_taslak_key]
                df = pd.DataFrame(URUNLER)
                arama = st.text_input(
                    "🔍 **Ürün Ara (Adı veya Kodu):**",
                    "",
                    key=f"urun_arama_{secilen_sube}_{bugun_str}",
                )
                filtre_df = (
                    df[
                        df["ADI"].str.contains(arama, case=False, na=False)
                        | df["KODU"].str.contains(arama, case=False, na=False)
                    ]
                    if arama
                    else df
                )

                st.subheader("📦 Stok ve Sipariş Girişi (Kasa)")

                for _, row in filtre_df.iterrows():
                    kod = row["KODU"]
                    urun_taslagi = siparis_taslagi.setdefault(
                        kod,
                        {"urun_adi": row["ADI"], "stok": "0", "siparis": 0.0, "not": ""},
                    )
                    varsayilan_stok_str = str(urun_taslagi.get("stok", "0"))
                    varsayilan_siparis = float(urun_taslagi.get("siparis", 0.0) or 0.0)

                    dolu_key = f"dolu_{secilen_sube}_{bugun_str}_{kod}"
                    stok_key = f"stok_{secilen_sube}_{bugun_str}_{kod}"
                    sip_key = f"sip_{secilen_sube}_{bugun_str}_{kod}"

                    # Widget ilk kez oluşturuluyorsa taslaktaki değerleri yükle.
                    if dolu_key not in st.session_state:
                        st.session_state[dolu_key] = varsayilan_stok_str == "Reyon Dolu"
                    if stok_key not in st.session_state:
                        try:
                            st.session_state[stok_key] = float(varsayilan_stok_str)
                        except (TypeError, ValueError):
                            st.session_state[stok_key] = 0.0
                    if sip_key not in st.session_state:
                        st.session_state[sip_key] = varsayilan_siparis

                    with st.expander(f"**{row['ADI']}** *(Kod: {kod})*"):
                        col1, col2 = st.columns([1.5, 1])
                        with col1:
                            stok_dolu = st.checkbox(
                                "🟢 Reyon Dolu (Depo Boş)",
                                key=dolu_key,
                            )
                            if not stok_dolu:
                                stok_val = st.number_input(
                                    "Mevcut Stok (Kasa)",
                                    min_value=0.0,
                                    step=1.0,
                                    key=stok_key,
                                )
                                stok_kayit = str(int(stok_val))
                            else:
                                stok_kayit = "Reyon Dolu"
                                st.caption("📌 *Stok 'Reyon Dolu' olarak kaydedilecek.*")

                        with col2:
                            siparis = st.number_input(
                                "Sipariş (Kasa)",
                                min_value=0.0,
                                step=1.0,
                                key=sip_key,
                            )
                        not_key = f"urun_not_{secilen_sube}_{bugun_str}_{kod}"
                        if not_key not in st.session_state:
                            st.session_state[not_key] = str(urun_taslagi.get("not", "") or "")
                        urun_notu = st.text_input(
                            "📝 Ürün Notu (isteğe bağlı)",
                            max_chars=150,
                            placeholder="Örn: Biraz sert olsun, yeşil gelsin...",
                            key=not_key,
                        )

                    # Görünen ürünün son değerlerini tam taslağa işle.
                    siparis_taslagi[kod] = {
                        "urun_adi": row["ADI"],
                        "stok": stok_kayit,
                        "siparis": float(siparis),
                        "not": str(urun_notu or "").strip(),
                    }

                genel_not_key = f"genel_siparis_notu_{secilen_sube}_{bugun_str}"
                if genel_not_key not in st.session_state:
                    st.session_state[genel_not_key] = bugun_genel_not
                genel_siparis_notu = st.text_area(
                    "📝 Genel Sipariş Notu (isteğe bağlı)",
                    max_chars=500,
                    placeholder="Siparişin tamamı için açıklama yazabilirsiniz...",
                    key=genel_not_key,
                )

                # Arama sonucu görünmeyen ürünler dahil tüm taslaktan kayıt listesi üret.
                kaydedilecek_veriler = []
                kaydedilecek_urun_notlari = {}
                for urun in URUNLER:
                    kod = urun["KODU"]
                    veri = siparis_taslagi.get(
                        kod, {"stok": "0", "siparis": 0.0, "urun_adi": urun["ADI"]}
                    )
                    stok_kayit = str(veri.get("stok", "0"))
                    try:
                        siparis_miktari = float(veri.get("siparis", 0.0) or 0.0)
                    except (TypeError, ValueError):
                        siparis_miktari = 0.0

                    urun_notu = str(veri.get("not", "") or "").strip()
                    if urun_notu:
                        kaydedilecek_urun_notlari[kod] = urun_notu

                    if stok_kayit != "0" or siparis_miktari > 0:
                        kaydedilecek_veriler.append({
                            "sube": secilen_sube,
                            "tarih": bugun_str,
                            "urun_kodu": kod,
                            "urun_adi": urun["ADI"],
                            "mevcut_stok": stok_kayit,
                            "siparis_miktari": siparis_miktari,
                        })

                st.divider()
                btn_col1, btn_col2, btn_col3 = st.columns([2, 1.15, 1.15])
                with btn_col1:
                    if st.button("💾 Siparişleri Güncelle / Kaydet", type="primary", use_container_width=True):
                        with st.spinner("Sipariş güvenli şekilde kaydediliyor..."):
                            sonuc = guvenli_sorgu(
                                "Sipariş kaydetme",
                                lambda: sube_siparisini_degistir(secilen_sube, bugun_str, kaydedilecek_veriler, st.session_state.get(siparis_snapshot_key), kullanici=secilen_sube)
                            )
                        if sonuc:
                            not_sonucu = guvenli_sorgu(
                                "Sipariş notlarını kaydetme",
                                lambda: siparis_notlarini_kaydet(
                                    secilen_sube, bugun_str, kaydedilecek_urun_notlari, genel_siparis_notu
                                ),
                            )
                            if not_sonucu is None:
                                st.warning("⚠️ Sipariş kaydedildi ancak notlar kaydedilemedi. Supabase'de siparis_notlari tablosunu kontrol edin.")
                            if kaydedilecek_veriler:
                                st.success(f"✅ **{secilen_sube}** şubesinin siparişi başarıyla kaydedildi!")
                                st.session_state[siparis_snapshot_key] = kayit_ozeti(kaydedilecek_veriler)
                            else:
                                st.warning("⚠️ Tüm değerler 0 olduğu için bugünkü sipariş temizlendi.")
                                st.session_state[siparis_snapshot_key] = kayit_ozeti([])
                            st.rerun()

                with btn_col2:
                    st.caption("Başka kullanıcı değişiklik yaptıysa:")
                    if st.button(
                        "🔄 Güncel Veriyi Yeniden Yükle",
                        type="secondary",
                        use_container_width=True,
                        help="Ekrandaki kaydedilmemiş taslağı temizler ve Supabase'deki en güncel siparişi yeniden açar.",
                    ):
                        siparis_oturumunu_temizle(secilen_sube, bugun_str)
                        st.toast("Güncel sipariş verileri yeniden yükleniyor...", icon="🔄")
                        st.rerun()

                with btn_col3:
                    iptal_onayi = st.checkbox("Sipariş iptalini onaylıyorum", key=f"iptal_onay_{secilen_sube}")
                    if st.button("🗑️ Bugünkü Siparişi İptal Et", type="secondary", use_container_width=True, disabled=not iptal_onayi):
                        sonuc = guvenli_sorgu(
                            "Sipariş iptali",
                            lambda: sube_siparisini_degistir(
                                secilen_sube, bugun_str, [], st.session_state.get(siparis_snapshot_key), kullanici=secilen_sube
                            )
                        )
                        if sonuc:
                            guvenli_sorgu(
                                "Sipariş notlarını silme",
                                lambda: supabase.table("siparis_notlari").delete().eq("sube", secilen_sube).eq("tarih", bugun_str).execute(),
                            )
                            siparis_oturumunu_temizle(secilen_sube, bugun_str)
                            st.success("🗑️ Bugünkü sipariş tamamen silindi.")
                            st.rerun()

                # ŞUBE SİPARİŞ DÖKÜMÜ VE GEÇMİŞ SİPARİŞLER
                st.divider()
                st.markdown("## 🧾 Sipariş Dökümü ve Geçmiş Siparişler")
                st.caption("Bugünkü veya geçmiş bir siparişi görüntüleyebilir, Excel indirebilir ya da doğrudan yazdırıp PDF olarak kaydedebilirsiniz.")

                gecmis_col1, gecmis_col2 = st.columns(2)
                with gecmis_col1:
                    baslangic_tarihi = st.date_input(
                        "Başlangıç Tarihi",
                        value=simdi_tr().date() - timedelta(days=30),
                        max_value=simdi_tr().date(),
                        key=f"gecmis_baslangic_{secilen_sube}",
                    )
                with gecmis_col2:
                    bitis_tarihi = st.date_input(
                        "Bitiş Tarihi",
                        value=simdi_tr().date(),
                        max_value=simdi_tr().date(),
                        key=f"gecmis_bitis_{secilen_sube}",
                    )

                if baslangic_tarihi > bitis_tarihi:
                    st.warning("⚠️ Başlangıç tarihi bitiş tarihinden sonra olamaz.")
                else:
                    gecmis_veriler = guvenli_veri_oku(
                        "Geçmiş siparişleri okuma",
                        lambda: supabase.table("siparisler")
                            .select("sube,tarih,urun_kodu,urun_adi,mevcut_stok,siparis_miktari")
                            .eq("sube", secilen_sube)
                            .gte("tarih", baslangic_tarihi.strftime("%Y-%m-%d"))
                            .lte("tarih", bitis_tarihi.strftime("%Y-%m-%d"))
                            .order("tarih", desc=True)
                            .execute().data or [],
                    )

                    gecmis_notlar = siparis_notlarini_oku(
                        secilen_sube,
                        baslangic_tarihi.strftime("%Y-%m-%d"),
                        bitis_tarihi.strftime("%Y-%m-%d"),
                    )
                    notlar_tarih_map = {}
                    for n in gecmis_notlar:
                        notlar_tarih_map.setdefault(str(n.get("tarih")), []).append(n)

                    tarih_gruplari = {}
                    for kayit in gecmis_veriler:
                        try:
                            miktar = float(kayit.get("siparis_miktari") or 0)
                        except (TypeError, ValueError):
                            miktar = 0.0
                        if miktar > 0:
                            tarih_gruplari.setdefault(str(kayit.get("tarih")), []).append(kayit)

                    if not tarih_gruplari:
                        st.info("ℹ️ Seçilen tarih aralığında sipariş kaydı bulunamadı.")
                    else:
                        ozet_satirlari = []
                        for tarih_key, kayitlar in tarih_gruplari.items():
                            toplam = sum(float(k.get("siparis_miktari") or 0) for k in kayitlar)
                            ozet_satirlari.append({
                                "Tarih": datetime.strptime(tarih_key, "%Y-%m-%d").strftime("%d.%m.%Y"),
                                "Ürün Çeşidi": len(kayitlar),
                                "Toplam Sipariş": f"{toplam:g} Kasa",
                            })
                        st.dataframe(pd.DataFrame(ozet_satirlari), use_container_width=True, hide_index=True)

                        tarih_secenekleri = sorted(tarih_gruplari.keys(), reverse=True)
                        secilen_gecmis_tarih = st.selectbox(
                            "📅 Dökümünü almak istediğiniz siparişi seçin",
                            tarih_secenekleri,
                            format_func=lambda x: datetime.strptime(x, "%Y-%m-%d").strftime("%d.%m.%Y"),
                            key=f"gecmis_tarih_sec_{secilen_sube}",
                        )
                        detay_kayitlari, detay_genel_not = notlari_kayitlara_ekle(
                            tarih_gruplari[secilen_gecmis_tarih],
                            notlar_tarih_map.get(secilen_gecmis_tarih, []),
                        )
                        detay_df = pd.DataFrame(detay_kayitlari)[["urun_kodu", "urun_adi", "mevcut_stok", "siparis_miktari", "urun_notu"]]
                        detay_df = detay_df.rename(columns={
                            "urun_kodu": "Ürün Kodu", "urun_adi": "Ürün Adı",
                            "mevcut_stok": "Mevcut Stok", "siparis_miktari": "Sipariş (Kasa)",
                            "urun_notu": "Ürün Notu",
                        })
                        st.dataframe(detay_df, use_container_width=True, hide_index=True)
                        if detay_genel_not:
                            st.info(f"📝 **Genel Sipariş Notu:** {detay_genel_not}")

                        excel_bytes = generate_sube_tek_siparis_excel(
                            secilen_sube, secilen_gecmis_tarih, detay_kayitlari, detay_genel_not
                        )
                        d_col1, d_col2 = st.columns(2)
                        with d_col1:
                            st.download_button(
                                "📊 EXCEL SİPARİŞ DÖKÜMÜNÜ İNDİR",
                                data=excel_bytes,
                                file_name=f"{secilen_sube}_{secilen_gecmis_tarih}_siparis.xlsx".replace(" ", "_"),
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary",
                                use_container_width=True,
                                key=f"excel_siparis_{secilen_sube}_{secilen_gecmis_tarih}",
                            )
                        with d_col2:
                            st.info("Aşağıdaki yeşil butona basınca yazdırma ekranı açılır. Yazıcı seçebilir veya **PDF olarak kaydet** seçeneğini kullanabilirsiniz.")

                        yazdir_html = generate_sube_siparis_html(secilen_sube, secilen_gecmis_tarih, detay_kayitlari, detay_genel_not)
                        components.html(yazdir_html, height=650, scrolling=True)

    # 2. HAL DAĞITIM PANELİ
    elif rol == "🚛 Hal Dağıtım Paneli":
        st.markdown("<h2 style='text-align: center;'>🚛 Hal Satınalma ve Dağıtım Paneli</h2>", unsafe_allow_html=True)

        if not st.session_state.hal_authed:
            hal_pin = st.text_input("🔑 Lütfen Satınalma/Hal Yetkili Şifresini Giriniz:", type="password")
            if st.button("Giriş Yap", type="primary"):
                if hal_pin == HAL_SIFRESI or hal_pin == YONETICI_SIFRESI:
                    st.session_state.hal_authed = True
                    st.rerun()
                else:
                    st.error("❌ Hatalı Satınalma Şifresi!")
        else:
            st.success("🔓 Satınalma Yetkili Girişi Aktif")
            if st.button("🔒 Oturumu Kapat"):
                st.session_state.hal_authed = False
                st.rerun()

            st.divider()
            st.markdown("#### 📅 Sevkiyat ve Dağıtım Tarihi Seçimi")
            t_col1, t_col2 = st.columns([2, 5])
            with t_col1:
                secilen_hal_tarihi = st.date_input("İşlem Yapmak İstediğiniz Tarih:", value=simdi_tr().date())
                hal_tarih_str = secilen_hal_tarihi.strftime('%Y-%m-%d')
            with t_col2:
                st.write("") 
                if secilen_hal_tarihi == simdi_tr().date():
                    st.info("🟢 **Bugünün** verileri ve dağıtım listesi görüntüleniyor.")
                else:
                    st.warning(f"🟡 **{secilen_hal_tarihi.strftime('%d.%m.%Y')}** tarihine ait dağıtım verileri görüntüleniyor.")

            st.divider()
            is_today = (secilen_hal_tarihi == simdi_tr().date())
            tarih_label = "BUGÜNÜN" if is_today else f"{secilen_hal_tarihi.strftime('%d.%m.%Y')} TARİHLİ"

            toplu_excel_bytes = generate_toplu_hal_excel(hal_tarih_str)
            if toplu_excel_bytes:
                st.download_button(
                    label=f"🚚 {tarih_label} TÜM SEVKİYAT DAĞITIM LİSTESİNİ İNDİR (YAZDIRMAYA HAZIR EXCEL)",
                    data=toplu_excel_bytes,
                    file_name=f"Toplu_Hal_Sevkiyat_Listesi_{hal_tarih_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
            else:
                st.caption(f"ℹ️ *{secilen_hal_tarihi.strftime('%d.%m.%Y')} tarihi için henüz halden girilmiş herhangi bir dağıtım bulunmuyor.*")

            st.divider()
            urun_listesi_adlar = [f"{u['ADI']} ({u['KODU']})" for u in URUNLER]
            secilen_urun_combo = st.selectbox("🛒 **Halden Alınan Ürünü Seçin:**", urun_listesi_adlar)
            secilen_urun_kod = secilen_urun_combo.split("(")[-1].replace(")", "").strip()
            secilen_urun_ad = secilen_urun_combo.split("(")[0].strip()

            hal_mevcut = guvenli_veri_oku(
                "Hal dağıtım kaydını okuma",
                lambda: supabase.table("hal_dagitim").select("sube,tarih,urun_kodu,urun_adi,dağıtılan_miktar").eq("tarih", hal_tarih_str).eq("urun_kodu", secilen_urun_kod).execute().data or []
            )
            hal_snapshot_key = f"hal_snapshot_{hal_tarih_str}_{secilen_urun_kod}"
            if hal_snapshot_key not in st.session_state:
                st.session_state[hal_snapshot_key] = kayit_ozeti(hal_mevcut)

            # Hal dağıtım alanları ürün ve tarih bazında ayrı taslaklarda tutulur.
            # Ürün değiştirilmeden hemen önce mevcut widget değerleri taslağa alınır;
            # geri dönüldüğünde yeni widget anahtarlarıyla bu değerler tekrar yüklenir.
            hal_taslaklari = st.session_state.setdefault("hal_taslaklari", {})
            hal_taslak_key = f"{hal_tarih_str}|{secilen_urun_kod}"
            yeni_baglam = hal_taslak_key
            onceki_baglam = st.session_state.get("hal_widget_baglam")
            onceki_anahtarlar = st.session_state.get("hal_widget_anahtarlar", {})

            # Önceki üründe girilmiş ama henüz kaydedilmemiş değerleri kaybetme.
            if onceki_baglam and onceki_anahtarlar:
                onceki_dagitim = {}
                for sube_adi in SUBE_LISTESI:
                    eski_key = onceki_anahtarlar.get("dagitim", {}).get(sube_adi)
                    onceki_dagitim[sube_adi] = float(st.session_state.get(eski_key, 0.0) or 0.0) if eski_key else 0.0
                eski_toplam_key = onceki_anahtarlar.get("toplam")
                hal_taslaklari[onceki_baglam] = {
                    "hal_toplam": float(st.session_state.get(eski_toplam_key, 0.0) or 0.0) if eski_toplam_key else 0.0,
                    "dagitim": onceki_dagitim,
                }

            kayitli_dagitim = {sube: 0.0 for sube in SUBE_LISTESI}
            for kayit in hal_mevcut:
                sube = kayit.get("sube")
                if sube in kayitli_dagitim:
                    try:
                        kayitli_dagitim[sube] = float(kayit.get("dağıtılan_miktar") or 0.0)
                    except (TypeError, ValueError):
                        kayitli_dagitim[sube] = 0.0

            kayitli_toplam = float(sum(kayitli_dagitim.values()))
            if hal_taslak_key not in hal_taslaklari:
                hal_taslaklari[hal_taslak_key] = {
                    "hal_toplam": kayitli_toplam,
                    "dagitim": kayitli_dagitim.copy(),
                }
            elif hal_mevcut:
                # Veritabanında kayıt varsa dağıtım değerlerini esas al; gerçek toplam alış
                # miktarı oturum taslağında daha yüksekse onu koru.
                mevcut_taslak = hal_taslaklari[hal_taslak_key]
                taslak_toplam = float(mevcut_taslak.get("hal_toplam", 0.0) or 0.0)
                hal_taslaklari[hal_taslak_key] = {
                    "hal_toplam": max(taslak_toplam, kayitli_toplam),
                    "dagitim": kayitli_dagitim.copy(),
                }

            # Her ürün/tarih geçişinde yeni widget anahtarları üretmek, Streamlit'in
            # önceki sıfır değerlerini kayıtlı verilerin üzerine yazmasını engeller.
            if onceki_baglam != yeni_baglam:
                st.session_state["hal_widget_surum"] = int(st.session_state.get("hal_widget_surum", 0)) + 1
            widget_surum = int(st.session_state.get("hal_widget_surum", 1))
            aktif_hal_taslagi = hal_taslaklari[hal_taslak_key]

            hal_toplam_widget_key = f"hal_toplam_{hal_tarih_str}_{secilen_urun_kod}_{widget_surum}"
            dagitim_widget_keys = {
                sube_adi: f"hal_dag_{hal_tarih_str}_{sube_adi}_{secilen_urun_kod}_{widget_surum}"
                for sube_adi in SUBE_LISTESI
            }

            if hal_toplam_widget_key not in st.session_state:
                st.session_state[hal_toplam_widget_key] = float(aktif_hal_taslagi.get("hal_toplam", 0.0) or 0.0)
            for sube_adi, widget_key in dagitim_widget_keys.items():
                if widget_key not in st.session_state:
                    st.session_state[widget_key] = float(
                        aktif_hal_taslagi.get("dagitim", {}).get(sube_adi, 0.0) or 0.0
                    )

            st.session_state["hal_widget_baglam"] = yeni_baglam
            st.session_state["hal_widget_anahtarlar"] = {
                "toplam": hal_toplam_widget_key,
                "dagitim": dagitim_widget_keys,
            }

            hal_toplam_kasa = st.number_input(
                f"📦 **Halden Alınan Toplam Miktar ({secilen_urun_ad}):**",
                min_value=0.0,
                step=1.0,
                key=hal_toplam_widget_key,
            )

            st.subheader("🏬 Şubelere Dağıtım Tablosu")
            dagitim_dict = {}
            toplam_dagitilan = 0.0
            d_col1, d_col2 = st.columns(2)
            for i, sube_adi in enumerate(SUBE_LISTESI):
                target_col = d_col1 if i % 2 == 0 else d_col2
                widget_key = dagitim_widget_keys[sube_adi]
                with target_col:
                    val = st.number_input(
                        f"📍 {sube_adi}:",
                        min_value=0.0,
                        step=1.0,
                        key=widget_key,
                    )
                    dagitim_dict[sube_adi] = float(val)
                    toplam_dagitilan += float(val)

            # Her yeniden çalıştırmada güncel alanları taslağa aktar.
            hal_taslaklari[hal_taslak_key] = {
                "hal_toplam": float(hal_toplam_kasa),
                "dagitim": {sube: float(miktar) for sube, miktar in dagitim_dict.items()},
            }

            kalan_kasa = hal_toplam_kasa - toplam_dagitilan
            st.divider()
            m1, m2, m3 = st.columns(3)
            m1.metric("Halden Alınan", f"{hal_toplam_kasa:.0f} Kasa")
            m2.metric("Şubelere Dağıtılan", f"{toplam_dagitilan:.0f} Kasa")
            if kalan_kasa < 0:
                m3.metric("⚠️ Fazla Dağıtılan", f"{abs(kalan_kasa):.0f} Kasa", delta_color="inverse")
                st.error("⚠️ Halden aldığınız miktardan daha fazla dağıtım yaptınız!")
            else:
                m3.metric("Kalan (Depo/Yedek)", f"{kalan_kasa:.0f} Kasa")

            st.divider()
            h_btn1, h_btn2 = st.columns(2)
            with h_btn1:
                if st.button("💾 Hal Dağıtımını Kaydet ve Şubelere Bildir", type="primary", use_container_width=True):
                    if hal_toplam_kasa == 0:
                        st.warning("⚠️ Halden alınan miktar 0 olamaz.")
                    elif kalan_kasa < 0:
                        st.error("❌ Hata: Alınan miktardan fazlası dağıtılamaz!")
                    else:
                        kayit_listesi = []
                        for sube, miktar in dagitim_dict.items():
                            if miktar > 0:
                                kayit_listesi.append({
                                    "sube": sube,
                                    "tarih": hal_tarih_str,
                                    "urun_kodu": secilen_urun_kod,
                                    "urun_adi": secilen_urun_ad,
                                    "dağıtılan_miktar": float(miktar)
                                })
                        if len(kayit_listesi) > 0:
                            with st.spinner("Dağıtım kaydı güvenli şekilde güncelleniyor..."):
                                sonuc = guvenli_sorgu(
                                    "Hal dağıtımı kaydetme",
                                    lambda: hal_dagitimini_degistir(hal_tarih_str, secilen_urun_kod, kayit_listesi, st.session_state.get(hal_snapshot_key), kullanici="Hal Yetkilisi")
                                )
                            if sonuc:
                                st.success(f"✅ **{secilen_urun_ad}** dağıtımı kaydedildi/güncellendi!")
                                guncel_hal = guvenli_veri_oku(
                                    "Kaydedilen hal dağıtımını doğrulama",
                                    lambda: supabase.table("hal_dagitim").select(
                                        "sube,tarih,urun_kodu,urun_adi,dağıtılan_miktar"
                                    ).eq("tarih", hal_tarih_str).eq("urun_kodu", secilen_urun_kod).execute().data or [],
                                    varsayilan=kayit_listesi,
                                )
                                st.session_state[hal_snapshot_key] = kayit_ozeti(guncel_hal)
                                # Başarılı kayıttan sonra taslağı doğrudan kaydedilen değerlerle güncelle.
                                hal_taslaklari[hal_taslak_key] = {
                                    "hal_toplam": float(hal_toplam_kasa),
                                    "dagitim": {sube: float(miktar) for sube, miktar in dagitim_dict.items()},
                                }
                                st.rerun()
                        else:
                            st.warning("⚠️ Şubelere herhangi bir miktar girilmedi.")
            with h_btn2:
                if toplam_dagitilan > 0 and kalan_kasa >= 0:
                    hal_excel_bytes = generate_hal_excel(secilen_urun_ad, secilen_urun_kod, hal_toplam_kasa, dagitim_dict, kalan_kasa, hal_tarih_str)
                    st.download_button(
                        label="📄 Sadece Bu Ürünün Excel Listesini İndir",
                        data=hal_excel_bytes,
                        file_name=f"Hal_Dagitim_{secilen_urun_kod}_{hal_tarih_str}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

    # 3. MERKEZ YÖNETİM PANELİ
    elif rol == "👑 Merkez Panel":
        st.markdown("<h2 style='text-align: center;'>🔒 Merkez Yönetim Paneli</h2>", unsafe_allow_html=True)

        if not st.session_state.admin_authed:
            sifre_giris = st.text_input("🔑 Lütfen Yönetim Şifresini Giriniz:", type="password")
            if st.button("Giriş Yap", type="primary"):
                if sifre_giris == YONETICI_SIFRESI:
                    st.session_state.admin_authed = True
                    st.rerun()
                else:
                    st.error("❌ Hatalı şifre!")
        else:
            st.success("🔓 Yetkili Girişi Başarılı")
            if st.button("🚪 Oturumu Kapat"):
                st.session_state.admin_authed = False
                st.rerun()

            f_col1, f_col2, f_col3 = st.columns([1.2, 1, 2])
            with f_col1:
                secilen_tarih = st.date_input("📅 Tarih Seçin", value=simdi_tr().date())
                tarih_str = secilen_tarih.strftime('%Y-%m-%d')
            with f_col2:
                filtre_sube = st.selectbox("🏬 Şube Filtresi", ["Tümü"] + SUBE_LISTESI)
            with f_col3:
                arama_admin = st.text_input("🔍 Ürün Arama (Ad / Kod)", "")

            st.divider()

            # SEKMELER
            tab_dashboard, tab_sip, tab_hal, tab_log, tab_yonetim = st.tabs([
                "📊 Yönetim Dashboard",
                "🛒 Şube Sipariş ve Stok Matrisi", 
                "🚛 Hal Sevkiyat ve Dağıtım Verileri",
                "🧾 İşlem Geçmişi",
                "🗑️ Veri / Geçmiş Yönetimi (Silme)"
            ])

            with tab_dashboard:
                dash_sip = guvenli_veri_oku(
                    "Dashboard sipariş verilerini okuma",
                    lambda: supabase.table("siparisler").select("sube,urun_kodu,urun_adi,siparis_miktari").eq("tarih", tarih_str).execute().data or [],
                )
                dash_hal = guvenli_veri_oku(
                    "Dashboard hal verilerini okuma",
                    lambda: supabase.table("hal_dagitim").select("sube,urun_kodu,urun_adi,dağıtılan_miktar").eq("tarih", tarih_str).execute().data or [],
                )

                df_ds = pd.DataFrame(dash_sip)
                df_dh = pd.DataFrame(dash_hal)
                if not df_ds.empty:
                    df_ds["siparis_miktari"] = pd.to_numeric(df_ds["siparis_miktari"], errors="coerce").fillna(0)
                if not df_dh.empty:
                    df_dh["dağıtılan_miktar"] = pd.to_numeric(df_dh["dağıtılan_miktar"], errors="coerce").fillna(0)

                siparis_toplam = float(df_ds["siparis_miktari"].sum()) if not df_ds.empty else 0
                sevk_toplam = float(df_dh["dağıtılan_miktar"].sum()) if not df_dh.empty else 0
                giren_subeler = sorted(df_ds["sube"].dropna().unique().tolist()) if not df_ds.empty else []
                urun_sayisi = int(df_ds.loc[df_ds["siparis_miktari"] > 0, "urun_kodu"].nunique()) if not df_ds.empty else 0

                st.subheader(f"📊 {secilen_tarih.strftime('%d.%m.%Y')} Yönetim Özeti")
                k1, k2, k3, k4 = st.columns(4)
                k1.metric("Sipariş Giren Şube", f"{len(giren_subeler)} / {len(SUBE_LISTESI)}")
                k2.metric("Sipariş Verilen Ürün", f"{urun_sayisi}")
                k3.metric("Toplam Sipariş", f"{siparis_toplam:.0f} Kasa")
                k4.metric("Toplam Sevkiyat", f"{sevk_toplam:.0f} Kasa")

                eksik_subeler = [s for s in SUBE_LISTESI if s not in giren_subeler]
                durum_col1, durum_col2 = st.columns(2)
                with durum_col1:
                    st.markdown("#### 🟢 Sipariş Giren Şubeler")
                    st.write(", ".join(giren_subeler) if giren_subeler else "Henüz sipariş giren şube yok.")
                with durum_col2:
                    st.markdown("#### 🔴 Sipariş Girmeyen Şubeler")
                    st.write(", ".join(eksik_subeler) if eksik_subeler else "Tüm şubeler siparişini girdi.")

                if not df_ds.empty and siparis_toplam > 0:
                    st.divider()
                    g1, g2 = st.columns(2)
                    with g1:
                        st.markdown("#### 🏬 Şube Bazlı Toplam Sipariş")
                        sube_grafik = df_ds.groupby("sube", as_index=False)["siparis_miktari"].sum().sort_values("siparis_miktari", ascending=False)
                        st.bar_chart(sube_grafik.set_index("sube"))
                    with g2:
                        st.markdown("#### 🥇 En Çok Sipariş Verilen 10 Ürün")
                        urun_grafik = df_ds.groupby("urun_adi", as_index=False)["siparis_miktari"].sum().sort_values("siparis_miktari", ascending=False).head(10)
                        st.bar_chart(urun_grafik.set_index("urun_adi"))

                    top_urunler = df_ds.groupby(["urun_kodu", "urun_adi"], as_index=False)["siparis_miktari"].sum().sort_values("siparis_miktari", ascending=False).head(10)
                    top_urunler = top_urunler.rename(columns={"urun_kodu":"Ürün Kodu", "urun_adi":"Ürün Adı", "siparis_miktari":"Toplam Sipariş (Kasa)"})
                    st.dataframe(top_urunler, use_container_width=True, hide_index=True)
                else:
                    st.info("ℹ️ Seçilen tarihte dashboard oluşturacak sipariş verisi bulunmuyor.")

                if not df_dh.empty and sevk_toplam > 0:
                    st.divider()
                    st.markdown("#### 🚚 Şube Bazlı Sevkiyat")
                    sevk_grafik = df_dh.groupby("sube", as_index=False)["dağıtılan_miktar"].sum().sort_values("dağıtılan_miktar", ascending=False)
                    st.bar_chart(sevk_grafik.set_index("sube"))

            # SEKME 1: ÇOK SÜTUNLU YAN YANA STOK VE SİPARİŞ MATRİSİ
            with tab_sip:
                query = supabase.table("siparisler").select("*").eq("tarih", tarih_str)
                if filtre_sube != "Tümü":
                    query = query.eq("sube", filtre_sube)
                siparis_merkez_verileri = guvenli_veri_oku(
                    "Merkez sipariş verilerini okuma",
                    lambda: query.execute().data or [],
                )

                if siparis_merkez_verileri:
                    df_res = pd.DataFrame(siparis_merkez_verileri)
                    df_res['siparis_miktari'] = pd.to_numeric(df_res['siparis_miktari'], errors='coerce').fillna(0)
                    df_res['mevcut_stok'] = df_res['mevcut_stok'].fillna("0").astype(str)

                    siparis_veren_subeler = sorted(df_res['sube'].dropna().unique().tolist())
                    siparis_vermeyen_subeler = [s for s in SUBE_LISTESI if s not in siparis_veren_subeler]
                    tamamlanma = int(round((len(siparis_veren_subeler) / len(SUBE_LISTESI)) * 100))
                    d1, d2, d3 = st.columns(3)
                    d1.metric("✅ Sipariş Giren", f"{len(siparis_veren_subeler)} / {len(SUBE_LISTESI)} Şube")
                    d2.metric("⏳ Sipariş Girmeyen", f"{len(siparis_vermeyen_subeler)} Şube")
                    d3.metric("📈 Tamamlanma", f"%{tamamlanma}")
                    with st.expander("🏬 Şube Sipariş Durumu", expanded=False):
                        for s_name in SUBE_LISTESI:
                            st.write(("🟢" if s_name in siparis_veren_subeler else "🔴") + f" {s_name}")

                    if arama_admin:
                        df_res = df_res[
                            df_res['urun_adi'].str.contains(arama_admin, case=False, na=False) | 
                            df_res['urun_kodu'].str.contains(arama_admin, case=False, na=False)
                        ]

                    if not df_res.empty:
                        m1, m2, m3 = st.columns(3)
                        m1.metric("📦 Toplam Sipariş Veren Şube", f"{df_res['sube'].nunique()} Şube")
                        m2.metric("🍉 Toplam Sipariş Kalemi", f"{len(df_res)} Kalem")
                        m3.metric("📊 Toplam Şube Sipariş Miktarı", f"{int(df_res['siparis_miktari'].sum())} Kasa")

                        st.subheader("📊 Şube Bazlı Stok ve Sipariş Matrisi")

                        # Seçilen tarihe ait ürün ve genel sipariş notlarını oku.
                        # Notlar doğrudan bu matriste gösterilir; satın alma görevlisinin
                        # Hal Dağıtım Paneli'ne geçmesine gerek kalmaz.
                        merkez_not_kayitlari = guvenli_veri_oku(
                            "Merkez sipariş notlarını okuma",
                            lambda: supabase.table("siparis_notlari")
                                .select("sube,tarih,urun_kodu,urun_notu,genel_not")
                                .eq("tarih", tarih_str)
                                .execute().data or [],
                            varsayilan=[],
                        )

                        urun_not_haritasi = {}
                        genel_not_haritasi = {}
                        for not_kaydi in merkez_not_kayitlari:
                            not_sube = str(not_kaydi.get("sube") or "").strip()
                            not_kodu = str(not_kaydi.get("urun_kodu") or "").strip()
                            urun_notu = str(not_kaydi.get("urun_notu") or "").strip()
                            genel_not = str(not_kaydi.get("genel_not") or "").strip()
                            if not_kodu and urun_notu:
                                urun_not_haritasi.setdefault(not_kodu, []).append({
                                    "Şube": not_sube,
                                    "Ürün Notu": urun_notu,
                                })
                            if not_sube and genel_not:
                                genel_not_haritasi[not_sube] = genel_not

                        # Geniş matris verisini üretelim
                        unique_urunler = df_res[['urun_kodu', 'urun_adi']].drop_duplicates().values
                        matrix_rows = []

                        for kod, adi in unique_urunler:
                            row_data = {"urun_kodu": kod, "urun_adi": adi}
                            t_sip_sum = 0
                            stok_adet_list = []
                            rd_sayisi = 0

                            for s_name in SUBE_LISTESI:
                                match_item = df_res[(df_res['urun_kodu'] == kod) & (df_res['sube'] == s_name)]
                                if not match_item.empty:
                                    stk = str(match_item.iloc[0]['mevcut_stok'])
                                    sip = float(match_item.iloc[0]['siparis_miktari'])
                                    row_data[f"{s_name}_stok"] = stk
                                    row_data[f"{s_name}_sip"] = sip if sip > 0 else "-"
                                    t_sip_sum += sip
                                    
                                    if stk == "Reyon Dolu":
                                        rd_sayisi += 1
                                    else:
                                        try:
                                            stok_adet_list.append(float(stk))
                                        except (ValueError, TypeError):
                                            pass
                                else:
                                    row_data[f"{s_name}_stok"] = "-"
                                    row_data[f"{s_name}_sip"] = "-"

                            row_data["toplam_sip"] = t_sip_sum
                            not_adedi = len(urun_not_haritasi.get(str(kod), []))
                            row_data["notlar"] = f"🟡 {not_adedi} Not" if not_adedi else "-"
                            toplam_stok = int(sum(stok_adet_list)) if stok_adet_list else 0
                            if rd_sayisi > 0:
                                row_data["toplam_stok"] = f"{toplam_stok} Kasa (+{rd_sayisi} RD)"
                            else:
                                row_data["toplam_stok"] = f"{toplam_stok} Kasa"

                            matrix_rows.append(row_data)

                        df_wide = pd.DataFrame(matrix_rows)

                        # Streamlit Multi-Index gösterimi için kolonları hazırlayalım
                        columns_tuples = [("Ürün Kodu", ""), ("Ürün Adı", "")]
                        for s_name in SUBE_LISTESI:
                            columns_tuples.append((s_name, "Stok"))
                            columns_tuples.append((s_name, "Sip."))
                        columns_tuples.append(("GENEL TOPLAM", "Top. Stok / RD"))
                        columns_tuples.append(("GENEL TOPLAM", "Top. Sipariş"))
                        columns_tuples.append(("📝 NOTLAR", "Adet"))

                        df_display = pd.DataFrame()
                        df_display[("Ürün Kodu", "")] = df_wide["urun_kodu"]
                        df_display[("Ürün Adı", "")] = df_wide["urun_adi"]

                        for s_name in SUBE_LISTESI:
                            df_display[(s_name, "Stok")] = df_wide[f"{s_name}_stok"]
                            df_display[(s_name, "Sip.")] = df_wide[f"{s_name}_sip"]

                        df_display[("GENEL TOPLAM", "Top. Stok / RD")] = df_wide["toplam_stok"]
                        df_display[("GENEL TOPLAM", "Top. Sipariş")] = df_wide["toplam_sip"]
                        df_display[("📝 NOTLAR", "Adet")] = df_wide["notlar"]

                        df_display.columns = pd.MultiIndex.from_tuples(columns_tuples)
                        st.dataframe(df_display, use_container_width=True, hide_index=True)

                        notlu_urunler = [
                            (str(kod), str(adi))
                            for kod, adi in unique_urunler
                            if urun_not_haritasi.get(str(kod))
                        ]
                        toplam_urun_notu = sum(len(v) for v in urun_not_haritasi.values())
                        toplam_genel_not = len(genel_not_haritasi)

                        with st.expander(
                            f"📝 Şube Sipariş Notları — {toplam_urun_notu} ürün notu, {toplam_genel_not} genel not",
                            expanded=bool(toplam_urun_notu or toplam_genel_not),
                        ):
                            if notlu_urunler:
                                not_urun_secimi = st.selectbox(
                                    "Notlarını görmek istediğiniz ürünü seçin",
                                    options=[f"{adi} ({kod})" for kod, adi in notlu_urunler],
                                    key=f"merkez_not_urun_{tarih_str}",
                                )
                                not_urun_kodu = not_urun_secimi.rsplit("(", 1)[-1].rstrip(")").strip()
                                secilen_notlar = urun_not_haritasi.get(not_urun_kodu, [])
                                if secilen_notlar:
                                    st.dataframe(
                                        pd.DataFrame(secilen_notlar),
                                        use_container_width=True,
                                        hide_index=True,
                                    )
                            else:
                                st.info("ℹ️ Seçilen tarihte ürün bazlı şube notu bulunmuyor.")

                            if genel_not_haritasi:
                                st.markdown("#### 📌 Şubelerin Genel Sipariş Notları")
                                st.dataframe(
                                    pd.DataFrame([
                                        {"Şube": sube_adi, "Genel Sipariş Notu": not_metni}
                                        for sube_adi, not_metni in genel_not_haritasi.items()
                                    ]),
                                    use_container_width=True,
                                    hide_index=True,
                                )

                        st.divider()

                        st.markdown("#### 📄 Rapor ve Yazdırma Merkezi")
                        st.caption("Yazdırma raporu A4 dikey biçimde, en fazla 2 sayfa ve onaylanan şube sırasıyla hazırlanır. Hücreler Stok/Sipariş formatındadır; RD, Reyon Dolu anlamına gelir.")

                        excel_duzenleme = generate_sube_siparis_excel(tarih_str, df_wide, "standart")
                        excel_dikey = generate_sube_siparis_dikey_2_sayfa(tarih_str, df_wide)

                        rapor_col1, rapor_col2 = st.columns(2)
                        with rapor_col1:
                            st.download_button(
                                label="📊 Excel İndir (Düzenleme)",
                                data=excel_duzenleme,
                                file_name=f"Sube_Stok_Siparis_Duzenleme_{tarih_str}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                        with rapor_col2:
                            st.download_button(
                                label="🖨️ A4 Dikey Yazdırma (2 Sayfa)",
                                data=excel_dikey,
                                file_name=f"Sube_Stok_Siparis_Dikey_2_Sayfa_{tarih_str}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary",
                                use_container_width=True
                            )

                    else:
                        st.info("ℹ️ Aranan kriterlere uygun şube sipariş verisi bulunamadı.")
                else:
                    st.info(f"ℹ️ {secilen_tarih.strftime('%d.%m.%Y')} tarihi için henüz kaydedilmiş bir şube siparişi bulunmuyor.")

            # SEKME 2: HAL DAĞITIM
            with tab_hal:
                query_h = supabase.table("hal_dagitim").select("*").eq("tarih", tarih_str)
                if filtre_sube != "Tümü":
                    query_h = query_h.eq("sube", filtre_sube)
                hal_merkez_verileri = guvenli_veri_oku(
                    "Merkez hal sevkiyat verilerini okuma",
                    lambda: query_h.execute().data or [],
                )

                if hal_merkez_verileri:
                    df_h = pd.DataFrame(hal_merkez_verileri)
                    df_h['dağıtılan_miktar'] = pd.to_numeric(df_h['dağıtılan_miktar'], errors='coerce').fillna(0)
                    if arama_admin:
                        df_h = df_h[
                            df_h['urun_adi'].str.contains(arama_admin, case=False, na=False) | 
                            df_h['urun_kodu'].str.contains(arama_admin, case=False, na=False)
                        ]
                    if not df_h.empty:
                        st.subheader("🚚 Halden Şubelere Sevk Edilen Mal Dağıtım Matrisi")
                        pivot_h = pd.pivot_table(df_h, values='dağıtılan_miktar', index=['urun_kodu', 'urun_adi'], columns=['sube'], aggfunc='sum', fill_value=0)
                        pivot_h['TOPLAM SEVK'] = pivot_h.sum(axis=1)
                        st.dataframe(pivot_h, use_container_width=True)

                        excel_bytes = generate_toplu_hal_excel(tarih_str)
                        if excel_bytes:
                            st.download_button(
                                label="📥 Seçilen Tarihin Hal Sevkiyat Raporunu Excel Olarak İndir",
                                data=excel_bytes,
                                file_name=f"Hal_Sevkiyat_Raporu_{tarih_str}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary",
                                use_container_width=True
                            )
                    else:
                        st.info("ℹ️ Aranan kriterlere uygun hal sevkiyat verisi bulunamadı.")
                else:
                    st.info(f"ℹ️ {secilen_tarih.strftime('%d.%m.%Y')} tarihi için halden yapılmış bir dağıtım/sevkiyat kaydı bulunmuyor.")


            # SEKME 4: İŞLEM GEÇMİŞİ
            with tab_log:
                st.subheader("🧾 İşlem Geçmişi")
                st.caption("Şube siparişi ve hal dağıtımı üzerindeki ekleme, güncelleme ve silme hareketleri burada görünür.")
                loglar = guvenli_veri_oku(
                    "İşlem geçmişini okuma",
                    lambda: supabase.table("islem_loglari").select("*").eq("kayit_tarihi", tarih_str).order("islem_zamani", desc=True).limit(1000).execute().data or [],
                )
                if loglar:
                    df_log = pd.DataFrame(loglar)
                    gosterilecek = [c for c in ["islem_zamani", "kullanici", "rol", "sube", "islem", "urun_kodu", "urun_adi", "eski_deger", "yeni_deger", "detay"] if c in df_log.columns]
                    df_log = df_log[gosterilecek]
                    if "islem_zamani" in df_log.columns:
                        # Supabase timestamptz değerlerini UTC olarak döndürebilir.
                        # Önce UTC olarak yorumlayıp ardından Türkiye saatine çeviriyoruz.
                        df_log["islem_zamani"] = (
                            pd.to_datetime(df_log["islem_zamani"], errors="coerce", utc=True)
                            .dt.tz_convert("Europe/Istanbul")
                            .dt.strftime("%d.%m.%Y %H:%M:%S")
                        )
                    df_log = df_log.rename(columns={
                        "islem_zamani":"İşlem Zamanı", "kullanici":"Kullanıcı", "rol":"Rol", "sube":"Şube",
                        "islem":"İşlem", "urun_kodu":"Ürün Kodu", "urun_adi":"Ürün Adı",
                        "eski_deger":"Eski Değer", "yeni_deger":"Yeni Değer", "detay":"Detay"
                    })
                    st.dataframe(df_log, use_container_width=True, hide_index=True)
                    st.download_button(
                        "📥 İşlem Geçmişini CSV İndir",
                        data=df_log.to_csv(index=False).encode("utf-8-sig"),
                        file_name=f"Islem_Gecmisi_{tarih_str}.csv",
                        mime="text/csv",
                        use_container_width=True,
                    )
                else:
                    st.info("ℹ️ Bu tarihte işlem kaydı bulunamadı. islem_loglari tablosu henüz oluşturulmadıysa aşağıdaki SQL dosyasını Supabase'te bir kez çalıştırın.")

            # SEKME 3: GEÇMİŞ VERİLERİ SİLME / TEMİZLEME YÖNETİMİ
            with tab_yonetim:
                st.subheader("💾 Sistem Yedeği")
                st.caption("Siparişler ve hal dağıtım kayıtlarının tamamını tek Excel dosyası olarak indirir.")

                try:
                    yedek_bytes, siparis_yedek_sayisi, hal_yedek_sayisi, yedek_zamani = generate_tum_veri_yedegi()
                    yedek_dosya_adi = f"YalcinMarket_Tam_Yedek_{yedek_zamani.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"
                    y1, y2, y3 = st.columns(3)
                    y1.metric("Şube Sipariş Kayıtları", f"{siparis_yedek_sayisi:,}".replace(",", "."))
                    y2.metric("Hal Dağıtım Kayıtları", f"{hal_yedek_sayisi:,}".replace(",", "."))
                    y3.metric("Toplam Kayıt", f"{siparis_yedek_sayisi + hal_yedek_sayisi:,}".replace(",", "."))
                    st.download_button(
                        label="💾 TÜM VERİLERİ TEK EXCEL DOSYASI OLARAK YEDEKLE",
                        data=yedek_bytes,
                        file_name=yedek_dosya_adi,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True,
                    )
                    st.caption(f"Yedek hazırlanma zamanı: {yedek_zamani.strftime('%d.%m.%Y %H:%M:%S')}")
                except Exception as e:
                    st.error(f"❌ Yedek hazırlanamadı: {e}")

                st.divider()
                st.subheader("🗑️ Geçmiş Veri ve Kayıt Temizleme Paneli")
                st.warning("⚠️ Bu ekrandan seçtiğiniz tarihe ait verileri tamamen silebilirsiniz. Bu işlem geri alınamaz!")

                silme_tarihi = st.date_input("Silmek İstediğiniz Tarihi Seçin:", value=simdi_tr().date(), key="del_date_picker")
                silme_tarih_str = silme_tarihi.strftime('%Y-%m-%d')

                d_secim = st.radio("Hangi Tablodaki Verileri Temizlemek İstiyorsunuz?", ["Şube Siparişleri Tablosu", "Hal Dağıtım Tablosu", "Her İki Tabloyu da Temizle"])

                st.markdown("<br>", unsafe_allow_html=True)
                kalici_sil_onayi = st.checkbox(
                    f"{silme_tarihi.strftime('%d.%m.%Y')} tarihli verilerin kalıcı silinmesini onaylıyorum",
                    key="kalici_sil_onayi"
                )
                silme_metni = st.text_input(
                    "Onaylamak için SİL yazın:",
                    key="kalici_sil_yazi_onayi"
                )
                silme_hazir = kalici_sil_onayi and silme_metni.strip().upper() == "SİL"
                if st.button("🔥 SEÇİLEN TARİHİN VERİLERİNİ KALICI OLARAK SİL", type="primary", disabled=not silme_hazir):
                    try:
                        if d_secim in ["Şube Siparişleri Tablosu", "Her İki Tabloyu da Temizle"]:
                            supabase.table("siparisler").delete().eq("tarih", silme_tarih_str).execute()
                        if d_secim in ["Hal Dağıtım Tablosu", "Her İki Tabloyu da Temizle"]:
                            supabase.table("hal_dagitim").delete().eq("tarih", silme_tarih_str).execute()
                        islem_logu_yaz(
                            "Merkez Yönetici", "Merkez", "", "Tarih bazlı kalıcı veri silme",
                            "siparisler/hal_dagitim", silme_tarih_str, detay=d_secim
                        )
                        
                        st.success(f"✅ {silme_tarihi.strftime('%d.%m.%Y')} tarihine ait seçilen veriler başarıyla temizlendi!")
                        st.rerun()
                    except Exception as e:
                        st.error(f"❌ Veriler silinirken hata oluştu: {e}")
